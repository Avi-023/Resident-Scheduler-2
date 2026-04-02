# app.py
import io, json, re, math, random, time, pickle, os
from datetime import date, timedelta
from collections import Counter, defaultdict
from pathlib import Path

import streamlit as st
import pandas as pd
import numpy as np
import altair as alt

# --------------------------
# Data persistence for calendar sharing
# --------------------------
SCHEDULE_CACHE_FILE = Path(__file__).parent / ".schedule_cache.pkl"
CASE_LOG_IMPORT_FOLDER = Path(__file__).parent / "imports" / "case_logs"
CASE_LOG_CACHE_FILE = Path(__file__).parent / ".case_logs_cache.pkl"

def save_case_logs_to_cache(case_logs: dict):
    """Save case logs to file for persistence."""
    try:
        with open(CASE_LOG_CACHE_FILE, "wb") as f:
            pickle.dump({"case_logs": case_logs, "saved_at": time.time()}, f)
    except Exception as e:
        pass  # Fail silently

def load_case_logs_from_cache():
    """Load case logs from cache file."""
    try:
        if CASE_LOG_CACHE_FILE.exists():
            with open(CASE_LOG_CACHE_FILE, "rb") as f:
                data = pickle.load(f)
                return data.get("case_logs", {})
    except Exception:
        pass
    return {}

def scan_import_folder_for_case_logs():
    """Scan the import folder for new case log files and import them."""
    if not CASE_LOG_IMPORT_FOLDER.exists():
        CASE_LOG_IMPORT_FOLDER.mkdir(parents=True, exist_ok=True)
        return []

    imported_files = []
    processed_marker = CASE_LOG_IMPORT_FOLDER / ".processed"

    # Load list of already processed files
    processed_files = set()
    if processed_marker.exists():
        processed_files = set(processed_marker.read_text().strip().split("\n"))

    # Scan for new files
    for f in CASE_LOG_IMPORT_FOLDER.iterdir():
        if f.is_file() and f.suffix.lower() in [".xlsx", ".xls", ".csv"] and f.name not in processed_files:
            try:
                # Read the file
                if f.suffix.lower() == ".csv":
                    df = pd.read_csv(f)
                else:
                    df = pd.read_excel(f)

                if not df.empty:
                    imported_files.append((f.name, df))

                    # Mark as processed
                    processed_files.add(f.name)

            except Exception as e:
                pass  # Skip files that can't be read

    # Update processed marker
    if imported_files:
        processed_marker.write_text("\n".join(processed_files))

    return imported_files

def save_schedule_to_cache(dailies: dict, schedule_df, roster_df):
    """Save schedule data to file for access by calendar page."""
    try:
        data = {
            "dailies": dailies,
            "schedule_df": schedule_df,
            "roster_df": roster_df,
            "saved_at": time.time()
        }
        with open(SCHEDULE_CACHE_FILE, "wb") as f:
            pickle.dump(data, f)
    except Exception as e:
        st.warning(f"Could not save schedule cache: {e}")

def load_schedule_from_cache():
    """Load schedule data from file."""
    try:
        if SCHEDULE_CACHE_FILE.exists():
            with open(SCHEDULE_CACHE_FILE, "rb") as f:
                return pickle.load(f)
    except Exception:
        pass
    return None

# --------------------------
# Schedule Version Management
# --------------------------
def save_schedule_snapshot(name: str):
    """Save current schedule state as a named snapshot."""
    if "schedule_snapshots" not in st.session_state:
        st.session_state.schedule_snapshots = {}

    snapshot = {
        "name": name,
        "saved_at": time.time(),
        "roster_table": st.session_state.get("roster_table", pd.DataFrame()).copy(),
        "schedule_df": st.session_state.get("sched_df", pd.DataFrame()).copy(),
        "dailies": {k: v.copy() for k, v in st.session_state.get("dailies", {}).items()},
    }
    st.session_state.schedule_snapshots[name] = snapshot
    return snapshot

def load_schedule_snapshot(name: str):
    """Load a named snapshot into current state."""
    if "schedule_snapshots" not in st.session_state:
        return False
    if name not in st.session_state.schedule_snapshots:
        return False

    snapshot = st.session_state.schedule_snapshots[name]
    st.session_state.roster_table = snapshot["roster_table"].copy()
    st.session_state.sched_df = snapshot["schedule_df"].copy()
    st.session_state.dailies = {k: v.copy() for k, v in snapshot["dailies"].items()}
    return True

def export_snapshot_to_json(name: str) -> str:
    """Export a snapshot to JSON string for download."""
    if "schedule_snapshots" not in st.session_state:
        return None
    if name not in st.session_state.schedule_snapshots:
        return None

    snapshot = st.session_state.schedule_snapshots[name]

    # Convert DataFrames to JSON-serializable format
    export_data = {
        "name": snapshot["name"],
        "saved_at": snapshot["saved_at"],
        "roster_table": snapshot["roster_table"].to_dict(orient="records"),
        "schedule_df": snapshot["schedule_df"].to_dict(orient="split"),
        "dailies": {k: v.to_dict(orient="split") for k, v in snapshot["dailies"].items()},
    }
    return json.dumps(export_data, indent=2)

def import_snapshot_from_json(json_str: str) -> dict:
    """Import a snapshot from JSON string."""
    try:
        data = json.loads(json_str)

        # Reconstruct DataFrames
        roster_table = pd.DataFrame(data["roster_table"])

        schedule_data = data["schedule_df"]
        schedule_df = pd.DataFrame(
            schedule_data["data"],
            index=schedule_data["index"],
            columns=schedule_data["columns"]
        )

        dailies = {}
        for k, v in data["dailies"].items():
            dailies[k] = pd.DataFrame(v["data"], index=v["index"], columns=v["columns"])

        snapshot = {
            "name": data["name"],
            "saved_at": data.get("saved_at", time.time()),
            "roster_table": roster_table,
            "schedule_df": schedule_df,
            "dailies": dailies,
        }
        return snapshot
    except Exception as e:
        return None

def get_snapshot_summary(snapshot: dict) -> dict:
    """Get summary stats for a snapshot."""
    schedule_df = snapshot.get("schedule_df", pd.DataFrame())
    roster_df = snapshot.get("roster_table", pd.DataFrame())

    if schedule_df.empty:
        return {"residents": 0, "blocks": 0}

    return {
        "residents": len(schedule_df),
        "blocks": len([c for c in schedule_df.columns if c.startswith("Block")]),
        "saved_at": snapshot.get("saved_at", 0),
    }

# --------------------------
# Optional libraries
# --------------------------
try:
    import pyarrow as _pa  # noqa: F401
    HAS_ARROW = True
except Exception:
    HAS_ARROW = False

try:
    import pulp  # Lightweight ILP backend if available
    HAS_PULP = True
except Exception:
    HAS_PULP = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except Exception:
    HAS_PDFPLUMBER = False

try:
    import pytesseract
    from pdf2image import convert_from_bytes
    from PIL import Image
    HAS_OCR = True
except Exception:
    HAS_OCR = False

def parse_pdf_with_ocr(file_buffer, debug=False):
    """Parse image-based PDF using OCR (Tesseract)."""
    if not HAS_OCR:
        if debug:
            st.error("OCR libraries not available (pytesseract, pdf2image)")
        return None

    try:
        file_buffer.seek(0)
        pdf_bytes = file_buffer.read()

        if debug:
            st.info("Converting PDF pages to images...")

        # Convert PDF pages to images
        images = convert_from_bytes(pdf_bytes, dpi=300)

        if debug:
            st.write(f"Converted {len(images)} pages to images")

        all_text = ""
        for i, image in enumerate(images):
            if debug:
                st.write(f"Running OCR on page {i+1}...")

            # Run OCR on each page
            page_text = pytesseract.image_to_string(image)
            all_text += page_text + "\n"

            if debug and page_text.strip():
                st.markdown(f"**Page {i+1} OCR text (first 1000 chars):**")
                st.code(page_text[:1000])

        if debug:
            st.write(f"Total OCR text length: {len(all_text)}")

        if all_text.strip():
            # Parse the OCR text
            df = parse_acgme_text(all_text, debug)
            return df

        return None

    except Exception as e:
        if debug:
            st.error(f"OCR error: {e}")
        return None

def parse_acgme_pdf(file_buffer, debug=False):
    """Parse ACGME case log PDF (Resident Minimum Defined Categories report)."""
    if not HAS_PDFPLUMBER:
        if debug:
            st.error("pdfplumber not installed")
        return None

    try:
        file_buffer.seek(0)
        all_rows = []
        all_text = ""

        with pdfplumber.open(file_buffer) as pdf:
            for page in pdf.pages:
                # Try table extraction first
                tables = page.extract_tables()
                for table in tables:
                    if table:
                        all_rows.extend(table)

                # Also get text for fallback
                page_text = page.extract_text()
                if page_text:
                    all_text += page_text + "\n"

        # Try table-based parsing first
        if all_rows:
            df = parse_pdf_table_rows(all_rows, debug)
            if df is not None and not df.empty:
                return df

        # Fall back to text-based parsing
        if all_text:
            df = parse_acgme_text(all_text, debug)
            if df is not None and not df.empty:
                return df

        # If no text was extracted, try OCR
        if len(all_text.strip()) == 0:
            if debug:
                st.warning("No text found in PDF. Trying OCR...")
            file_buffer.seek(0)
            return parse_pdf_with_ocr(file_buffer, debug)

        if debug:
            st.error(f"Text found but parsing failed. Text length: {len(all_text)}")
        return None

    except Exception as e:
        if debug:
            st.error(f"PDF parse error: {e}")
        return None

def parse_pdf_table_rows(all_rows, debug=False):
    """Parse rows extracted from PDF tables."""
    if not all_rows:
        return None

    # Find header row (contains "Category" or "Minimum" or "Defined")
    header_idx = 0
    for i, row in enumerate(all_rows):
        if row and any(cell and any(kw in str(cell).lower() for kw in ['category', 'minimum', 'defined']) for cell in row if cell):
            header_idx = i
            break

    if header_idx >= len(all_rows):
        return None

    headers = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1:]

    # Clean headers - remove None/empty, strip whitespace
    headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(headers)]

    # Filter out empty rows
    data_rows = [row for row in data_rows if row and any(cell for cell in row)]

    if not data_rows:
        return None

    # Ensure all rows have same length as headers
    cleaned_rows = []
    for row in data_rows:
        if len(row) < len(headers):
            row = list(row) + [None] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[:len(headers)]
        cleaned_rows.append(row)

    df = pd.DataFrame(cleaned_rows, columns=headers)

    if debug:
        st.write(f"Parsed {len(df)} rows with columns: {list(df.columns)}")

    return df

def parse_acgme_text(text_content, debug=False):
    """Parse ACGME text content when table extraction fails."""
    if not text_content:
        return None

    lines = text_content.strip().split('\n')
    data = []

    # First pass: find the header line to identify resident columns
    header_line = None
    resident_names = []
    for line in lines:
        if 'minimum' in line.lower() or 'category' in line.lower():
            header_line = line
            # Extract resident names (words after "Minimum" that aren't numbers)
            parts = re.split(r'\s{2,}', line)
            for p in parts:
                p = p.strip()
                if p and not p.replace(',', '').isdigit() and p.lower() not in ['category', 'minimum', 'defined', 'categories']:
                    # Likely a resident name
                    if len(p) > 2 and ' ' in p:  # Names usually have spaces
                        resident_names.append(p)
            break

    if debug:
        st.write(f"Found {len(lines)} lines, resident names: {resident_names}")

    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue

        # Skip header-like lines
        if 'minimum' in line_stripped.lower() and 'category' in line_stripped.lower():
            continue
        if line_stripped.lower().startswith('page '):
            continue

        # Try to parse lines with format: "Category Name    Minimum    Count1    Count2..."
        # Split by multiple spaces or tabs
        parts = re.split(r'\s{2,}|\t+', line_stripped)

        if len(parts) >= 2:
            # Separate text parts from number parts
            nums = []
            text_parts = []

            for p in parts:
                p = p.strip()
                # Check if it's a number (possibly with commas)
                clean_p = p.replace(',', '').replace('.', '')
                if clean_p.isdigit():
                    nums.append(int(p.replace(',', '')))
                elif p and not clean_p.isdigit():
                    text_parts.append(p)

            if nums and text_parts:
                category = ' '.join(text_parts)
                # Skip total/summary rows
                if 'total' in category.lower() and 'major' not in category.lower():
                    continue

                row_data = {'Category': category}
                if len(nums) >= 1:
                    row_data['Minimum'] = nums[0]
                if len(nums) >= 2:
                    # Additional columns are resident counts
                    if resident_names:
                        for i, name in enumerate(resident_names):
                            if i + 1 < len(nums):
                                row_data[name] = nums[i + 1]
                    else:
                        row_data['Count'] = nums[1]

                data.append(row_data)

    if data:
        df = pd.DataFrame(data)
        if debug:
            st.write(f"Text parsing found {len(df)} rows")
        return df

    return None

def _old_parse_acgme_text(text_content):
    """Legacy parser - kept for reference."""
    if not text_content:
        return None

    lines = text_content.strip().split('\n')
    data = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        parts = re.split(r'\s{2,}', line)
        if len(parts) >= 2:
            try:
                nums = []
                text_parts = []
                for p in parts:
                    p = p.strip()
                    if p.replace(',', '').isdigit():
                        nums.append(int(p.replace(',', '')))
                    else:
                        text_parts.append(p)

                if nums and text_parts:
                    category = ' '.join(text_parts)
                    data.append({
                        'Category': category,
                        'Minimum': nums[0] if len(nums) > 0 else 0,
                        'Count': nums[1] if len(nums) > 1 else nums[0]
                    })
            except:
                pass

    if data:
        return pd.DataFrame(data)
    return None

# --------------------------
# Colors & styling
# --------------------------
PALETTE = {
    "gold":       "#FFD966",
    "red_green":  "#008000",  # Red/Green merged team shown in GREEN
    "vascular":   "#8B0000",  # dark red
    "breast":     "#FFB6C1",  # light pink
    "chief":      "#000080",  # navy
    "icu":        "#FFFF00",  # yellow
    "floor":      "#D2B48C",
    "nights":     "#000000",  # black
    "elective":   "#D9B3FF",
    "pittsburgh": "#FFE599",
    "vacation":   "#B6D7A8",
}
WHITE_TEXT = {"nights", "chief", "vascular", "red_green"}
BOLD_TEXT  = {"icu", "vascular", "red_green"}

# Dropdown options for rotation assignments (Yearly tab)
ROTATION_OPTIONS = [
    "",  # Allow blank/empty
    "Gold",
    "Red/Green",
    "Vascular",
    "Breast",
    "Chief",
    "ICU",
    "Floor",
    "Nights",
    "Elective",
    "Pittsburgh",
    "Vacation",
]

# Dropdown options for daily assignments (Daily Blocks tab)
DAILY_OPTIONS = [
    "",  # Allow blank/empty
    "Gold",
    "Red/Green",
    "Vascular",
    "Breast",
    "Chief",
    "ICU",
    "Floor",
    "Nights",
    "Elective",
    "Pittsburgh",
    "Vacation",
    "Personal Day",
    "Interview",
    "ATLS",
    "Conference",
    "Surgical Jeopardy",
    "Cadaver Lab",
    "Superservice",
    "F/Su",
    "Sa",
]

# --------------------------
# Attending Physicians & Teams
# --------------------------
TEAMS = {
    "Gold": ["Dr. Morrissey", "Dr. Sleet", "Dr. Dumire", "Dr. Silvis"],
    "Red/Green": ["Dr. Duke", "Dr. Curfman"],
    "Vascular": ["Dr. Tretter", "Dr. Dekornfeld"],
    "Breast": ["Dr. Arlow"],
}

# Other specialties (not primary teams, but available for manual assignment)
OTHER_SPECIALTIES = {
    "Cardiothoracic": ["Dr. Sherwal", "Dr. Mavridis"],
    "Plastic Surgery": ["Dr. Shayesteh", "Dr. Rollins"],
    "Urology": ["Dr. Chason"],
}

ALL_ATTENDINGS = {}
for team, docs in TEAMS.items():
    for doc in docs:
        ALL_ATTENDINGS[doc] = team
for specialty, docs in OTHER_SPECIALTIES.items():
    for doc in docs:
        ALL_ATTENDINGS[doc] = specialty

# --------------------------
# Clinic Coverage Rules
# --------------------------
CLINIC_RULES = {
    "Gold": {
        "covers_clinic": True,
        "min_residents": 1,
        "max_residents": 3,
        "pgy5_half_day_only": True,
        "cross_cover_from": ["Red/Green"],  # Can borrow from Red if needed
    },
    "Red/Green": {
        "covers_clinic": True,
        "min_residents": 1,
        "max_residents": 3,
        "pgy5_half_day_only": True,
        "cross_cover_from": ["Gold"],  # Can borrow from Gold if needed
    },
    "Vascular": {
        "covers_clinic": False,  # Residents do NOT cover vascular clinic
    },
    "Breast": {
        "covers_clinic": True,
        "min_residents": 1,
        "max_residents": 1,
        "daily_except_or": True,  # Every day except operative days
        "pgy5_half_day_only": True,
    },
}

# Attending-specific preferences
ATTENDING_PREFERENCES = {
    "Dr. Curfman": {
        "clinic_residents_preferred": 2,  # Prefers 2-3 residents
        "clinic_residents_max": 3,
    },
}

# --------------------------
# ACGME General Surgery Defined Categories & Minimums
# Based on ACGME/ABS requirements
# --------------------------
# ACGME General Surgery Defined Categories (matches ACGME Case Log System export format)
ACGME_CATEGORIES = {
    "Skin and Soft Tissue": {
        "minimum": 25,
        "subcategories": {}
    },
    "Breast": {
        "minimum": 40,
        "subcategories": {
            "Mastectomy": {"minimum": 5},
            "Axilla": {"minimum": 5},
        }
    },
    "Head and Neck": {
        "minimum": 25,
        "subcategories": {}
    },
    "Alimentary Tract": {
        "minimum": 180,
        "subcategories": {
            "Esophagus": {"minimum": 5},
            "Stomach": {"minimum": 10},
            "Small Intestine": {"minimum": 20},
            "Large Intestine": {"minimum": 40},
            "Appendix": {"minimum": 30},
            "Anorectal": {"minimum": 20},
            "Liver": {"minimum": 5},
            "Biliary": {"minimum": 35},
            "Pancreas": {"minimum": 3},
            "Spleen": {"minimum": 3},
        }
    },
    "Abdomen": {
        "minimum": 85,
        "subcategories": {
            "Abdominal - Other": {"minimum": 0},
            "Hernia - Inguinal": {"minimum": 40},
            "Hernia - Femoral": {"minimum": 0},
            "Hernia - Ventral/Incisional": {"minimum": 20},
            "Hernia - Other Abdominal Wall": {"minimum": 0},
        }
    },
    "Endocrine": {
        "minimum": 15,
        "subcategories": {
            "Thyroid": {"minimum": 5},
            "Parathyroid": {"minimum": 3},
            "Adrenal": {"minimum": 0},
        }
    },
    "Vascular": {
        "minimum": 25,
        "subcategories": {
            "Carotid": {"minimum": 0},
            "Abdominal Aorta": {"minimum": 0},
            "Visceral Vessels": {"minimum": 0},
            "Lower Extremity Bypass": {"minimum": 0},
            "Amputation": {"minimum": 0},
            "Vascular Access for Dialysis": {"minimum": 10},
            "Vein": {"minimum": 0},
        }
    },
    "Pediatric": {
        "minimum": 15,
        "subcategories": {}
    },
    "Thoracic": {
        "minimum": 20,
        "subcategories": {
            "Chest Wall": {"minimum": 0},
            "Lung/Pleura": {"minimum": 0},
            "Mediastinum": {"minimum": 0},
        }
    },
    "Trauma": {
        "minimum": 75,
        "subcategories": {
            "Operative Trauma": {"minimum": 20},
            "Non-operative Trauma": {"minimum": 40},
        }
    },
    "Critical Care": {
        "minimum": 100,
        "subcategories": {}
    },
    "Basic Laparoscopic": {
        "minimum": 100,
        "subcategories": {}
    },
    "Complex Laparoscopic": {
        "minimum": 50,
        "subcategories": {}
    },
    "Endoscopy": {
        "minimum": 85,
        "subcategories": {
            "Upper GI Endoscopy": {"minimum": 35},
            "Lower GI Endoscopy": {"minimum": 50},
        }
    },
}

# Total case minimum
ACGME_TOTAL_MINIMUM = 850

def parse_acgme_summary_report(df):
    """
    Parse ACGME summary report format (ResMinimumDefCat export).

    The format has columns like:
    - Category (with indented subcategories)
    - Minimum
    - One or more resident name columns with their totals

    Returns: dict of {resident_name: {category: {"total": N, "subcategories": {...}}}}
    """
    result = {}

    if df is None or df.empty:
        return result

    # Clean column names
    df.columns = df.columns.str.strip()

    # Find the category column (usually first column or "Category")
    cat_col = None
    for col in df.columns:
        if 'category' in col.lower() or col == df.columns[0]:
            cat_col = col
            break

    if cat_col is None:
        cat_col = df.columns[0]

    # Find minimum column
    min_col = None
    for col in df.columns:
        if 'minimum' in col.lower() or 'min' in col.lower():
            min_col = col
            break

    # Resident columns are any columns after category/minimum that contain numbers
    resident_cols = []
    for col in df.columns:
        if col != cat_col and col != min_col:
            # Check if this column has numeric data
            try:
                if df[col].dropna().apply(lambda x: str(x).replace(',', '').isdigit() if pd.notna(x) else True).any():
                    resident_cols.append(col)
            except:
                pass

    # If no obvious resident columns found, take remaining columns
    if not resident_cols:
        resident_cols = [c for c in df.columns if c not in [cat_col, min_col]]

    # Initialize result for each resident
    for res_col in resident_cols:
        result[res_col] = {}

    # Parse each row
    current_category = None
    for _, row in df.iterrows():
        cat_value = str(row.get(cat_col, "")).strip()

        if not cat_value or cat_value.lower() in ['nan', 'none', '']:
            continue

        # Detect if this is a subcategory (indented or has specific markers)
        is_subcategory = cat_value.startswith('    ') or cat_value.startswith('\t')
        cat_name = cat_value.strip()

        # Skip total rows
        if 'total' in cat_name.lower() and 'major' in cat_name.lower():
            continue
        if cat_name.lower() == 'total':
            continue

        # Get minimum value
        minimum = 0
        if min_col:
            try:
                min_val = row.get(min_col, 0)
                minimum = int(float(str(min_val).replace(',', ''))) if pd.notna(min_val) else 0
            except:
                pass

        # Process each resident's value
        for res_col in resident_cols:
            try:
                val = row.get(res_col, 0)
                count = int(float(str(val).replace(',', ''))) if pd.notna(val) else 0
            except:
                count = 0

            if is_subcategory and current_category:
                # Add as subcategory under current category
                if current_category not in result[res_col]:
                    result[res_col][current_category] = {"total": 0, "minimum": 0, "subcategories": {}}
                result[res_col][current_category]["subcategories"][cat_name] = {
                    "count": count,
                    "minimum": minimum
                }
            else:
                # Main category
                current_category = cat_name
                if cat_name not in result[res_col]:
                    result[res_col][cat_name] = {"total": 0, "minimum": minimum, "subcategories": {}}
                result[res_col][cat_name]["total"] = count
                result[res_col][cat_name]["minimum"] = minimum

    return result

# Common CPT code to category mappings (subset - kept for manual entry)
CPT_TO_CATEGORY = {
    # Alimentary Tract - Appendix
    "44950": ("ALIMENTARY TRACT", "Appendix"),  # Appendectomy
    "44960": ("ALIMENTARY TRACT", "Appendix"),  # Appendectomy with abscess
    "44970": ("ALIMENTARY TRACT", "Appendix"),  # Lap appendectomy

    # Alimentary Tract - Biliary
    "47562": ("ALIMENTARY TRACT", "Biliary"),   # Lap cholecystectomy
    "47563": ("ALIMENTARY TRACT", "Biliary"),   # Lap cholecystectomy with cholangiography
    "47600": ("ALIMENTARY TRACT", "Biliary"),   # Cholecystectomy
    "47610": ("ALIMENTARY TRACT", "Biliary"),   # Cholecystectomy with exploration

    # Alimentary Tract - Large Intestine
    "44140": ("ALIMENTARY TRACT", "Large Intestine"),  # Colectomy, partial
    "44141": ("ALIMENTARY TRACT", "Large Intestine"),  # Colectomy with colostomy
    "44143": ("ALIMENTARY TRACT", "Large Intestine"),  # Colectomy with anastomosis
    "44204": ("ALIMENTARY TRACT", "Large Intestine"),  # Lap colectomy, partial
    "44207": ("ALIMENTARY TRACT", "Large Intestine"),  # Lap colectomy with anastomosis

    # Alimentary Tract - Small Intestine
    "44120": ("ALIMENTARY TRACT", "Small Intestine"),  # Enterectomy
    "44121": ("ALIMENTARY TRACT", "Small Intestine"),  # Enterectomy, additional

    # Alimentary Tract - Anorectal
    "46255": ("ALIMENTARY TRACT", "Anorectal"),  # Hemorrhoidectomy
    "46260": ("ALIMENTARY TRACT", "Anorectal"),  # Hemorrhoidectomy, complex
    "46270": ("ALIMENTARY TRACT", "Anorectal"),  # Fistulotomy

    # Hernia
    "49505": ("ABDOMEN", "Hernia - Inguinal"),    # Inguinal hernia repair
    "49507": ("ABDOMEN", "Hernia - Inguinal"),    # Inguinal hernia, incarcerated
    "49520": ("ABDOMEN", "Hernia - Inguinal"),    # Inguinal hernia, recurrent
    "49650": ("ABDOMEN", "Hernia - Inguinal"),    # Lap inguinal hernia repair
    "49560": ("ABDOMEN", "Hernia - Ventral/Incisional"),  # Incisional hernia repair
    "49565": ("ABDOMEN", "Hernia - Ventral/Incisional"),  # Incisional hernia, recurrent
    "49652": ("ABDOMEN", "Hernia - Ventral/Incisional"),  # Lap ventral hernia repair
    "49653": ("ABDOMEN", "Hernia - Ventral/Incisional"),  # Lap incisional hernia repair

    # Breast
    "19301": ("BREAST", "Biopsy/Partial Mastectomy"),  # Partial mastectomy
    "19302": ("BREAST", "Biopsy/Partial Mastectomy"),  # Partial mastectomy with lymph node
    "19303": ("BREAST", "Mastectomy"),            # Simple mastectomy
    "19307": ("BREAST", "Mastectomy"),            # Modified radical mastectomy
    "38525": ("BREAST", "Sentinel Node"),          # Sentinel node biopsy

    # Endocrine
    "60240": ("ENDOCRINE", "Thyroid"),            # Thyroidectomy
    "60500": ("ENDOCRINE", "Parathyroid"),        # Parathyroidectomy
    "60650": ("ENDOCRINE", "Adrenal"),            # Adrenalectomy, lap

    # Vascular
    "35301": ("VASCULAR", "Carotid"),             # Carotid endarterectomy
    "36830": ("VASCULAR", "Dialysis Access"),     # AV fistula creation
    "36832": ("VASCULAR", "Dialysis Access"),     # AV fistula revision
    "36831": ("VASCULAR", "Dialysis Access"),     # AV graft

    # Endoscopy
    "43239": ("ENDOSCOPY", "EGD"),                # EGD with biopsy
    "43235": ("ENDOSCOPY", "EGD"),                # EGD diagnostic
    "45378": ("ENDOSCOPY", "Colonoscopy"),        # Colonoscopy diagnostic
    "45380": ("ENDOSCOPY", "Colonoscopy"),        # Colonoscopy with biopsy
    "45385": ("ENDOSCOPY", "Colonoscopy"),        # Colonoscopy with polypectomy

    # Laparoscopic (also counted in primary category)
    "47562": ("LAPAROSCOPIC/MINIMALLY INVASIVE", "Lap Cholecystectomy"),
    "44970": ("LAPAROSCOPIC/MINIMALLY INVASIVE", "Lap Appendectomy"),
    "49650": ("LAPAROSCOPIC/MINIMALLY INVASIVE", "Lap Hernia"),
    "44204": ("LAPAROSCOPIC/MINIMALLY INVASIVE", "Lap Colorectal"),
}

def norm_label(v: str) -> str:
    if not isinstance(v, str) or not v.strip():
        return ""
    t = v.strip().lower()
    if t == "gold": return "gold"
    if t in ("red/green","redgreen","rg","red-green"): return "red_green"
    if t.startswith("vasc"): return "vascular"
    if t.startswith("breast"): return "breast"
    if t in ("chief","trauma chief"): return "chief"
    if t in ("icu","trauma icu"): return "icu"
    if t in ("floor","trauma floor"): return "floor"
    if t == "nf" or t.startswith("night"): return "nights"
    if t.startswith("elective"): return "elective"
    if t.startswith("pitts"): return "pittsburgh"
    if t in ("liver","transplant","pancreas","pediatrics","peds"): return "pittsburgh"
    if t in ("vac","vacation","v"): return "vacation"
    return t

def style_cell(label: str) -> dict:
    key = norm_label(label)
    if not key or key not in PALETTE:
        return {}
    css = {"background-color": PALETTE[key]}
    css["color"] = "#FFFFFF" if key in WHITE_TEXT else "#000000"
    css["font-weight"] = "bold" if key in BOLD_TEXT else "normal"
    return css

def style_each_cell(v):
    css = style_cell(v)
    if not css: return ""
    return ";".join(f"{k}: {val}" for k, val in css.items())

# --------------------------
# Helpers
# --------------------------
WEEKDAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

def build_blocks(start: date, n_blocks: int = 13):
    return [(start + timedelta(days=28*i), start + timedelta(days=27 + 28*i)) for i in range(n_blocks)]

def hdr_for_block(s: date, e: date) -> str:
    return f"{s.month}/{s.day} - {e.month}/{e.day}"

def tab_name(s: date, e: date) -> str:
    return hdr_for_block(s, e)

def pgy_of(name, roster_map):
    return roster_map.get(name, {}).get("PGY")

def role_of(name, roster_map):
    p = pgy_of(name, roster_map)
    if p in ("PGY-5","PGY-4"): return "Senior"
    if p in ("PGY-3","PGY-2"): return "Junior"
    return "Intern"

def gini(arr):
    arr = np.array(arr, dtype=float)
    arr = arr[arr>=0]
    if arr.size == 0: return 0.0
    if np.all(arr==0): return 0.0
    arr = np.sort(arr)
    n = arr.size
    cum = np.cumsum(arr)
    return (n+1 - 2*np.sum(cum)/cum[-1]) / n

# Excel sheet-name sanitizer
_INVALID_EXCEL_CHARS = r'[:\\/?*\[\]]'
def safe_sheet_name(name: str, taken: set) -> str:
    n = re.sub(_INVALID_EXCEL_CHARS, '-', str(name)).strip().strip("'") or "Sheet"
    n = n[:31]
    base, i = n, 1
    while n in taken:
        suf = f" ({i})"
        n = base[:31-len(suf)] + suf
        i += 1
    taken.add(n)
    return n

# Pittsburgh helper: detect late-January overlap for the academic year
def block_has_late_january(s: date, e: date, ay_start: date) -> bool:
    january_year = ay_start.year + 1
    d = s
    while d <= e:
        if d.year == january_year and d.month == 1 and d.day >= 16:
            return True
        d += timedelta(days=1)
    return False

# --------------------------
# Lite-mode friendly table
# --------------------------
def show_table(df: pd.DataFrame, key: str, *, editable: bool, hide_index: bool=False,
               styler_subset=None, daily=False, index_name_hint="Resident",
               dropdown_columns=None, dropdown_options=None) -> pd.DataFrame:
    if HAS_ARROW:
        if editable:
            # For daily tables with tuple columns, flatten to strings for dropdown support
            if daily and isinstance(df.columns, pd.MultiIndex):
                # Flatten tuple columns to strings
                flat_df = df.copy()
                flat_df.columns = [f"{a}|{b}" for (a, b) in flat_df.columns]

                # Build column config for ALL flattened columns
                column_config = {}
                if dropdown_options:
                    for col in flat_df.columns:
                        column_config[col] = st.column_config.SelectboxColumn(
                            col,
                            options=dropdown_options,
                            required=False,
                        )

                edited_flat = st.data_editor(
                    flat_df,
                    use_container_width=True,
                    key=key,
                    hide_index=hide_index,
                    num_rows="dynamic",
                    column_config=column_config if column_config else None
                )

                # Convert back to MultiIndex columns
                tuples = []
                for c in edited_flat.columns:
                    if isinstance(c, str) and "|" in c:
                        a, b = c.split("|", 1)
                        tuples.append((a, b))
                    else:
                        tuples.append((c, ""))
                edited_flat.columns = pd.MultiIndex.from_tuples(tuples, names=["Weekday", "Date"])
                return edited_flat
            else:
                # Non-daily tables: use dropdown columns as specified
                column_config = {}
                if dropdown_columns and dropdown_options:
                    for col in dropdown_columns:
                        if col in df.columns:
                            column_config[col] = st.column_config.SelectboxColumn(
                                col,
                                options=dropdown_options,
                                required=False,
                            )
                return st.data_editor(
                    df,
                    use_container_width=True,
                    key=key,
                    hide_index=hide_index,
                    num_rows="dynamic",
                    column_config=column_config if column_config else None
                )
        else:
            if styler_subset is not None:
                st.dataframe(df.style.applymap(style_each_cell, subset=styler_subset),
                             use_container_width=True, hide_index=hide_index)
            else:
                st.dataframe(df, use_container_width=True, hide_index=hide_index)
            return df

    # Lite mode CSV fallback
    if editable:
        if daily:
            flat = df.copy()
            flat.columns = [f"{a}|{b}" for (a,b) in flat.columns]
            default_csv = flat.reset_index().to_csv(index=False)
            csv_text = st.text_area(f"CSV editor – {key}", default_csv, height=260, key=f"csv_{key}")
            try:
                tmp = pd.read_csv(io.StringIO(csv_text))
                if index_name_hint in tmp.columns:
                    tmp = tmp.set_index(index_name_hint)
                tuples = []
                for c in tmp.columns:
                    if isinstance(c, str) and "|" in c:
                        a, b = c.split("|", 1)
                        tuples.append((a, b))
                    else:
                        tuples.append((c, ""))
                tmp.columns = pd.MultiIndex.from_tuples(tuples, names=["Weekday","Date"])
                st.caption("Preview"); st.markdown(tmp.to_html(escape=False), unsafe_allow_html=True)
                return tmp
            except Exception as e:
                st.error(f"CSV parse error: {e}")
                st.markdown(df.to_html(escape=False), unsafe_allow_html=True)
                return df
        else:
            default_csv = df.reset_index().to_csv(index=False)
            csv_text = st.text_area(f"CSV editor – {key}", default_csv, height=240, key="csv_"+key)
            try:
                tmp = pd.read_csv(io.StringIO(csv_text))
                if index_name_hint in tmp.columns and (df.index.name in [None, index_name_hint]):
                    tmp = tmp.set_index(index_name_hint)
                st.caption("Preview"); st.markdown(tmp.to_html(escape=False), unsafe_allow_html=True)
                return tmp
            except Exception as e:
                st.error(f"CSV parse error: {e}")
                st.markdown(df.to_html(escape=False), unsafe_allow_html=True)
                return df
    else:
        if styler_subset is not None:
            st.markdown(df.style.applymap(style_each_cell, subset=styler_subset).to_html(), unsafe_allow_html=True)
        else:
            st.markdown(df.to_html(escape=False), unsafe_allow_html=True)
        return df

# --------------------------
# Roster normalization & utils
# --------------------------
def normalize_roster_input(df: pd.DataFrame) -> pd.DataFrame:
    import re as _re
    if df is None or df.empty:
        raise ValueError("Roster is empty.")
    df = df.copy()
    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame(df)

    # Auto-clean++: trim, drop blanks/"none"
    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].astype(str)

    col_norm = {c: str(c).strip().lower() for c in df.columns}
    inv = {v: k for k, v in col_norm.items()}

    def find_col(cands):
        for a in cands:
            if a in inv: return inv[a]
        return None

    resident_col = find_col(["resident","name","resident name"])
    if resident_col is None:
        if df.index.name and str(df.index.name).strip().lower() in {"resident","name","resident name"}:
            df = df.reset_index().rename(columns={df.columns[0]: "Resident"})
        elif df.index.dtype == "object" and df.index.notna().any():
            df = df.reset_index().rename(columns={df.columns[0]: "Resident"})
        else:
            obj_cols = [c for c in df.columns if df[c].dtype == "object"]
            if obj_cols:
                df = df.rename(columns={obj_cols[0]: "Resident"})
    else:
        if resident_col != "Resident":
            df = df.rename(columns={resident_col: "Resident"})

    # Standard columns
    pgy_col = find_col(["pgy","level"])
    if pgy_col and pgy_col != "PGY": df = df.rename(columns={pgy_col: "PGY"})
    prelim_col = find_col(["prelim","preliminary"])
    if prelim_col and prelim_col != "Prelim": df = df.rename(columns={prelim_col: "Prelim"})
    include_col = find_col(["include?","include","active"])
    if include_col and include_col != "Include?": df = df.rename(columns={include_col: "Include?"})

    # Optional attribute columns respected by engine
    for c in ["Notes","Vacation 1","Vacation 2","Vacation 3"]:
        if c not in df.columns: df[c] = ""

    # Ensure required cols
    if "Resident" not in df.columns:
        raise KeyError("Could not locate a 'Resident' column.")
    if "PGY" not in df.columns:
        raise KeyError("Could not locate a 'PGY' column.")
    if "Prelim" not in df.columns: df["Prelim"] = "N"
    if "Include?" not in df.columns: df["Include?"] = "Y"

    # Normalize
    def _norm_pgy(x: str) -> str:
        s = str(x).strip().upper()
        s = _re.sub(r"\(.*?\)", "", s)
        s = s.replace("_"," ").replace("-"," ")
        s = s.replace("YEAR","").replace("PGY","PGY ")
        s = _re.sub(r"\s+"," ", s).strip()
        m = _re.search(r"\b([1-5])\b", s)
        if m: return f"PGY-{m.group(1)}"
        return s

    def _yn(x):
        s = str(x).strip().lower()
        return "Y" if s in {"y","yes","true","1"} else ("N" if s in {"n","no","false","0"} else "Y")

    df["Resident"] = df["Resident"].astype(str)
    df["PGY"]      = df["PGY"].astype(str).map(_norm_pgy)
    df["Prelim"]   = df["Prelim"].astype(str).map(_yn)
    df["Include?"] = df["Include?"].astype(str).map(_yn)

    # Drop blanks/"none"
    bad = df["Resident"].str.strip().str.lower().isin(["", "none", "na", "n/a"])
    df = df[~bad].reset_index(drop=True)

    cols = ["Resident","PGY","Prelim","Include?","Notes","Vacation 1","Vacation 2","Vacation 3"]
    return df[[c for c in cols if c in df.columns]].reset_index(drop=True)

def validate_roster(df: pd.DataFrame) -> None:
    bad = []
    if df["Resident"].isna().any() or (df["Resident"].str.strip()=="").any():
        bad.append("Blank resident names")
    if df["Resident"].str.strip().str.lower().eq("none").any():
        bad.append("A row named 'None' (remove it)")
    dups = df["Resident"].str.strip().value_counts()
    dups = dups[dups > 1]
    if not dups.empty:
        bad.append(f"Duplicate names: {', '.join(dups.index.tolist())}")
    bad_pgy = ~df["PGY"].str.match(r"^PGY-[1-5]$")
    if bad_pgy.any():
        bad.append("Non-standard PGY values present (expect PGY-1..PGY-5)")
    if bad:
        raise ValueError("; ".join(bad))

def ensure_roster_size(df: pd.DataFrame, size: int) -> pd.DataFrame:
    df = df.copy()
    size = max(int(size), 1)
    if len(df) < size:
        add = size - len(df)
        add_df = pd.DataFrame({
            "Resident": ["" for _ in range(add)],
            "PGY": ["PGY-1"] * add,
            "Prelim": ["N"] * add,
            "Include?": ["Y"] * add,
            "Notes": ["" for _ in range(add)],
            "Vacation 1": ["" for _ in range(add)],
            "Vacation 2": ["" for _ in range(add)],
            "Vacation 3": ["" for _ in range(add)],
        })
        df = pd.concat([df, add_df], ignore_index=True)
    elif len(df) > size:
        df = df.iloc[:size].reset_index(drop=True)
    return df

def parse_csv_list(s: str):
    if not isinstance(s, str) or not s.strip():
        return []
    return [x.strip() for x in re.split(r"[;,]", s) if x.strip()]

# --------------------------
# Reason logging (explainability)
# --------------------------
def log_reason(reasons, res_name, block_hdr, why):
    reasons.append({"Resident": res_name, "Block": block_hdr, "Reason": why})

# --------------------------
# Auto-generate Yearly grid (Greedy)
# --------------------------
def auto_generate_yearly(roster_df: pd.DataFrame, start_date: date, constraints: dict, base_fixed: pd.DataFrame = None):
    roster_df = normalize_roster_input(roster_df)
    roster_df = roster_df[roster_df["Include?"].str.upper().eq("Y")].reset_index(drop=True)

    order_keys = {"PGY-5":0,"PGY-4":1,"PGY-3":2,"PGY-2":3,"PGY-1":4}
    roster_df["ord"] = roster_df["PGY"].map(order_keys)
    roster_df = roster_df.sort_values(["ord","Resident"]).drop(columns=["ord"]).reset_index(drop=True)

    names = list(roster_df["Resident"])
    roster_map = {r["Resident"]: {"PGY": r["PGY"], "Prelim": r.get("Prelim","N")} for _, r in roster_df.iterrows()}

    blocks = build_blocks(start_date, 13)
    headers = [hdr_for_block(s,e) for (s,e) in blocks]
    schedule_df = pd.DataFrame(index=names, columns=headers, dtype=object)

    # carry forward user-fixed entries
    if base_fixed is not None and not base_fixed.empty:
        base_fixed = base_fixed.reindex(index=names, columns=headers)
        mask = base_fixed.notna() & (base_fixed.apply(lambda col: col.astype(str).str.strip()) != "")
        schedule_df[mask] = base_fixed[mask]

    # Assignment reasons
    reasons = []

    # Helper: previous assignment in prior blocks
    def prev_assignment(name: str, bi: int):
        if bi <= 0: return None
        return schedule_df.loc[name, headers[bi-1]]

    def prev2_assignment(name: str, bi: int):
        if bi <= 1: return None
        return schedule_df.loc[name, headers[bi-2]]

    # Prevent >2 consecutive same rotation (service-agnostic helper; callers can override limit)
    def exceeds_consecutive_limit(name: str, bi: int, label: str, limit:int=2) -> bool:
        if bi < limit:
            return False
        last1 = prev_assignment(name, bi)
        last2 = prev2_assignment(name, bi)
        if last1 is None or last2 is None:
            return False
        return norm_label(last1) == norm_label(label) == norm_label(last2)

    # Helper: count current seniors on a team in this block
    def seniors_on(bi: int, label: str) -> int:
        lbl = label
        return sum(1 for n in names if schedule_df.loc[n, headers[bi]] == lbl and role_of(n, roster_map) == "Senior")

    # Caps (UI-cap values still available, but Gold/RG will be constrained to 3–4 by hard rule below)
    GOLD_CAP_UI = int(constraints.get("gold_cap", 4))
    RG_CAP_UI   = int(constraints.get("rg_cap", 4))
    VASC_MAX    = int(constraints.get("vascular_cap", 3))
    BREAST_MAX  = 1

    # Hard targets
    GOLD_MIN, GOLD_MAX = 3, 4
    RG_MIN,   RG_MAX   = 3, 4

    # Pittsburgh constraints
    ENABLE_PGH            = bool(constraints.get("enable_pittsburgh", True))
    PG3_ONLY_PGH          = bool(constraints.get("pg3_pittsburgh", True))
    PGH_BLOCK_BLACKLIST   = set(int(x) for x in constraints.get("pgh_block_blacklist", []))

    vasc_junior_counts = Counter()  # track PGY-2/3 used as vascular juniors
    vasc_intern_counts = Counter()  # track PGY-1 used as vascular interns

    # Gold/Red/Green fairness tracking by role
    gold_counts = {"Senior": Counter(), "Junior": Counter(), "Intern": Counter()}
    rg_counts = {"Senior": Counter(), "Junior": Counter(), "Intern": Counter()}

    # Night float tracking
    st.session_state.setdefault("_nf_counts", Counter())
    st.session_state.setdefault("_nf_role_counts", defaultdict(Counter))  # name -> {"Senior":x,"Junior":y,"Intern":z}
    st.session_state.setdefault("_last_nf", {})
    st.session_state.setdefault("nf_overrides", [])
    nf_counts = st.session_state["_nf_counts"]
    nf_role_counts = st.session_state["_nf_role_counts"]
    last_nf   = st.session_state["_last_nf"]

    MIN_NF_GAP_BLOCKS = 1

    # Service minima (kept for ICU; Gold/RG min enforced as hard rule)
    svc_min = constraints.get("service_minima", {"Gold":3, "Red/Green":3, "ICU":1})

    # Eligibility calendars (JSON spec)
    svc_unavail = constraints.get("eligibility_calendars", {})

    def is_unavailable(name, pgy, service_label, bi):
        key = service_label
        spec = svc_unavail.get(key, {})
        blist_pgy = spec.get(pgy, [])
        if bi in set(blist_pgy):
            return True
        byname = spec.get("names", {})
        if name in byname and bi in set(byname[name]):
            return True
        # PTO windows
        try:
            pto = next((x for x in roster_df.itertuples(index=False) if x.Resident==name), None)
            if pto and getattr(pto, "PTO Windows", ""):
                blocks_list = []
                for tok in parse_csv_list(getattr(pto, "PTO Windows")):
                    try: blocks_list.append(int(tok))
                    except Exception: pass
                if bi in set(blocks_list) and key not in {"Nights","Weekend Call"}:
                    return True
        except Exception:
            pass
        # Avoid Services (free-text list)
        try:
            row = next((x for x in roster_df.itertuples(index=False) if x.Resident==name), None)
            if row and getattr(row, "Avoid Services", ""):
                avoid_set = set(norm_label(s) for s in parse_csv_list(getattr(row, "Avoid Services")))
                if norm_label(key) in avoid_set:
                    return True
        except Exception:
            pass
        return False

    # Seniority helpers
    def is_senior(n): return role_of(n, roster_map) == "Senior"
    def is_junior(n): return role_of(n, roster_map) == "Junior"
    def is_intern(n): return roster_map[n]["PGY"] == "PGY-1"

    # prefer candidates who do NOT repeat last block on same service
    def repeat_penalty(name: str, bi: int, label: str) -> int:
        last = prev_assignment(name, bi)
        return 1 if (isinstance(last, str) and norm_label(last) == norm_label(label)) else 0

    def would_violate_senior_cap(name: str, bi: int, label: str) -> bool:
        return is_senior(name) and seniors_on(bi, label) >= 2

    def would_exceed_run(name: str, bi: int, label: str, limit:int=2) -> bool:
        return exceeds_consecutive_limit(name, bi, label, limit=limit)

    # Penalty: avoid two seniors from same PGY on the same team if a second senior is needed
    def senior_same_year_penalty(name: str, bi: int, label: str) -> int:
        if not is_senior(name): return 0
        existing = [x for x in names if schedule_df.loc[x, headers[bi]] == label and is_senior(x)]
        if len(existing) == 1:
            return 0 if pgy_of(name, roster_map) != pgy_of(existing[0], roster_map) else 1
        return 0

    def assign_first(pool, bi, label):
        """Fewest prior on that label; avoid consecutive repeat; forbid >2-run (or 3 for PGH); respect senior<=2; prefer 4/5 mix."""
        hdr = headers[bi]
        cand=[n for n in pool if pd.isna(schedule_df.loc[n, hdr])
              and not is_unavailable(n, roster_map[n]["PGY"], label, bi)
              and not would_violate_senior_cap(n, bi, label)
              and not would_exceed_run(n, bi, label, limit=(3 if label=="Pittsburgh" else 2))]
        if not cand: return None
        counts={n:int((schedule_df.loc[n]==label).sum()) for n in cand}
        pick=sorted(
            cand,
            key=lambda n: (
                counts[n],
                repeat_penalty(n, bi, label),
                senior_same_year_penalty(n, bi, label),
                rng.random(),  # Random tiebreaker for fairness
                n
            )
        )[0]
        schedule_df.loc[pick, headers[bi]]=label
        log_reason(reasons, pick, headers[bi], f"Assigned to {label} (min prior; avoid consecutive; prefer Sr mix)")
        return pick

    def team(bi,label): return [n for n in names if schedule_df.loc[n, headers[bi]]==label]
    def count_team_here(bi, label): return int((schedule_df[headers[bi]]==label).sum())

    def _nf_gap_ok(name: str, bi_: int, min_gap=MIN_NF_GAP_BLOCKS) -> bool:
        last = last_nf.get(name, None)
        return (last is None) or ((bi_ - last) >= (min_gap + 1))

    def _is_pgy5(name: str) -> int:
        return 1 if roster_map.get(name, {}).get("PGY") == "PGY-5" else 0

    rng = random.Random(int(constraints.get("random_seed", 0) or 0))

    # ========= Plan Pittsburgh 3-block runs (GLOBAL solve) =========
    if ENABLE_PGH:
        # Pre-filled Pittsburgh occupancy & per-resident prefilled indexes
        pre_by_res = defaultdict(set)
        preoccupied_blocks = set()
        for bi,_hdr in enumerate(headers):
            for n in names:
                if schedule_df.loc[n, headers[bi]] == "Pittsburgh":
                    pre_by_res[n].add(bi)
                    preoccupied_blocks.add(bi)

        # Locked rows: do not alter if they already have PGH; otherwise we still must try to place them
        locked_rows = set()
        for r in roster_df.itertuples(index=False):
            if getattr(r, "Lock", "").strip().lower() in {"row","true","y","yes"}:
                locked_rows.add(r.Resident)

        mandatory_pgh = [n for n in names if roster_map[n]["PGY"]=="PGY-3"]

        # Helper: block allowed under current policy
        def _block_allowed(t: int, honor_blacklist: bool) -> bool:
            # Block 1 (index 0) is never allowed for Pittsburgh
            if t == 0:
                return False
            if honor_blacklist and (t in PGH_BLOCK_BLACKLIST):
                return False
            s,e = blocks[t]
            # Always avoid late January
            if block_has_late_january(s,e,start_date):
                return False
            return True

        # Pre-preserve any existing exact 3-run
        fixed_windows = {}
        used_blocks = set(preoccupied_blocks)
        for n in mandatory_pgh:
            existing_sorted = sorted(pre_by_res.get(n, []))
            if not existing_sorted:
                continue
            runs=[]; start=None; prev=None
            for bi in existing_sorted:
                if start is None:
                    start=bi; prev=bi
                elif bi==prev+1:
                    prev=bi
                else:
                    runs.append((start,prev)); start=bi; prev=bi
            if start is not None: runs.append((start,prev))
            for a,b in runs:
                if (b-a+1)==3:
                    fixed_windows[n] = tuple(range(a,b+1))
                    used_blocks.update(range(a,b+1))
                    log_reason(reasons, n, headers[a], "Pittsburgh pre-fixed 3-block run preserved")
                    break  # take the first exact run

        def build_candidates(honor_blacklist: bool):
            cands = {}
            for n in mandatory_pgh:
                if n in fixed_windows:
                    continue
                lst=[]
                for s_idx in range(0, len(blocks)-2):
                    window = (s_idx, s_idx+1, s_idx+2)
                    if any(not _block_allowed(t, honor_blacklist) for t in window):
                        continue
                    # cannot overlap PGH blocks already preoccupied by other residents
                    if any((t in preoccupied_blocks and t not in pre_by_res.get(n,set())) for t in window):
                        continue
                    # name-level availability + not pre-filled with other non-PGH services
                    ok=True
                    for t in window:
                        hdr = headers[t]
                        if is_unavailable(n, roster_map[n]["PGY"], "Pittsburgh", t):
                            ok=False; break
                        v = schedule_df.loc[n, hdr]
                        if (pd.notna(v) and str(v).strip()!="" and v!="Pittsburgh"):
                            ok=False; break
                    if ok:
                        lst.append(window)
                cands[n]=lst
            return cands

        def solve_assignment(cands):
            # Backtracking over residents with least options first
            order = [n for n in mandatory_pgh if n not in fixed_windows]
            order.sort(key=lambda n: (len(cands.get(n, [])), n))
            assignment = dict(fixed_windows)  # start with fixed
            used = set(used_blocks)
            # mark used by fixed windows
            for w in fixed_windows.values():
                used.update(w)

            def dfs(i: int) -> bool:
                if i == len(order):
                    return True
                n = order[i]
                for w in cands.get(n, []):
                    if any(t in used for t in w):
                        continue
                    assignment[n] = w
                    used.update(w)
                    if dfs(i+1):
                        return True
                    # backtrack
                    used.difference_update(w)
                    assignment.pop(n, None)
                return False

            ok = dfs(0)
            return ok, assignment

        # Try with blacklist honored first
        cands = build_candidates(honor_blacklist=True)
        solved, assign_map = solve_assignment(cands)

        # If impossible, retry ignoring the blacklist (still avoids late Jan)
        if not solved:
            cands2 = build_candidates(honor_blacklist=False)
            solved, assign_map = solve_assignment(cands2)
            if solved:
                log_reason(reasons, "(PGH planner)", headers[0], "Replanned Pittsburgh ignoring blacklist to satisfy all PGY-3 runs")

        # Apply if solved
        if solved:
            for n, win in assign_map.items():
                for t in win:
                    schedule_df.loc[n, headers[t]] = "Pittsburgh"
                a = min(win); b = max(win)
                log_reason(reasons, n, headers[a], f"Pittsburgh 3-block run planned (blocks {a+1}-{b+1})")
        else:
            # Not solvable even without blacklist; leave for checks to flag
            log_reason(reasons, "(PGH planner)", headers[0], "Unable to schedule all PGY-3 Pittsburgh runs under hard constraints")

    # ========= Pre-plan Breast: ensure every PGY-2+ resident gets at least 1 block =========
    breast_eligible = [n for n in names if roster_map[n]["PGY"] in {"PGY-2", "PGY-3", "PGY-4", "PGY-5"}]
    breast_assigned = {n: 0 for n in breast_eligible}

    # Check for any pre-existing Breast assignments
    for n in breast_eligible:
        for hdr in headers:
            if schedule_df.loc[n, hdr] == "Breast":
                breast_assigned[n] += 1

    # Assign one Breast block to each PGY-2+ resident who doesn't have one yet
    # Shuffle to ensure fairness in block selection
    residents_needing_breast = [n for n in breast_eligible if breast_assigned[n] == 0]
    rng.shuffle(residents_needing_breast)

    for n in residents_needing_breast:
        # Find a block where this resident is not yet assigned and Breast is not taken
        available_blocks = []
        for bi, hdr in enumerate(headers):
            # Check if resident is unassigned this block and no one else has Breast
            if pd.isna(schedule_df.loc[n, hdr]) or str(schedule_df.loc[n, hdr]).strip() == "":
                breast_count_this_block = int((schedule_df[hdr] == "Breast").sum())
                if breast_count_this_block < BREAST_MAX:
                    available_blocks.append(bi)

        if available_blocks:
            # Pick the block where this resident has the fewest other commitments coming up
            chosen_bi = rng.choice(available_blocks)
            schedule_df.loc[n, headers[chosen_bi]] = "Breast"
            breast_assigned[n] += 1
            log_reason(reasons, n, headers[chosen_bi], "Breast pre-assigned (ensuring all PGY-2+ get at least 1)")

    # ---------------- Per-block assignment ----------------
    for bi,(bs,be) in enumerate(blocks):
        hdr = headers[bi]

        # Locked rows: skip assigning any locked rows (keep base_fixed or prior assignment)
        locked_rows = set()
        for r in roster_df.itertuples(index=False):
            if getattr(r, "Lock", "").strip().lower() in {"row","true","y","yes"}:
                locked_rows.add(r.Resident)

        # Chief: H1 PGY-3; H2 PGY-2
        if bi<6:
            chief=assign_first([n for n in names if roster_map[n]["PGY"]=="PGY-3" and n not in locked_rows], bi,"Chief")
            if chief is None:
                assign_first([n for n in names if roster_map[n]["PGY"]=="PGY-2" and n not in locked_rows], bi,"Chief")
        else:
            chief=assign_first([n for n in names if roster_map[n]["PGY"]=="PGY-2" and n not in locked_rows], bi,"Chief")
            if chief is None:
                assign_first([n for n in names if roster_map[n]["PGY"]=="PGY-3" and n not in locked_rows], bi,"Chief")

        # ICU (HARD): H1 PGY-2 only; H2 PGY-1 only; fair distribution; must be covered
        if bi < 6:
            elig = [n for n in names if roster_map[n]["PGY"]=="PGY-2" and pd.isna(schedule_df.loc[n, hdr])
                    and not is_unavailable(n, "PGY-2", "ICU", bi) and n not in locked_rows
                    and not would_exceed_run(n, bi, "ICU")]
            if elig:
                pick = sorted(elig, key=lambda n:(Counter(schedule_df.loc[n]== "ICU")[True], repeat_penalty(n, bi, "ICU"), n))[0]
                schedule_df.loc[pick, hdr] = "ICU"
                log_reason(reasons, pick, hdr, "ICU (H1 PGY-2 only) fair distribution (avoid consecutive)")
        else:
            elig = [n for n in names if roster_map[n]["PGY"]=="PGY-1" and pd.isna(schedule_df.loc[n, hdr])
                    and not is_unavailable(n, "PGY-1", "ICU", bi) and n not in locked_rows
                    and not would_exceed_run(n, bi, "ICU")]
            if elig:
                pick = sorted(elig, key=lambda n:(Counter(schedule_df.loc[n]== "ICU")[True], repeat_penalty(n, bi, "ICU"), n))[0]
                schedule_df.loc[pick, hdr] = "ICU"
                log_reason(reasons, pick, hdr, "ICU (H2 PGY-1 only) fair distribution (avoid consecutive)")

        # Night Float (HARD): must exist every block; Sr + Jr + Intern; ≥1-block gap preferred; avoid PGY-5 senior in last block
        def _pick_night_role(role: str, relax_gap: bool, avoid_pgy5_last: bool=False):
            pool = [n for n in names if pd.isna(schedule_df.loc[n, hdr]) and n not in locked_rows]
            pool = [n for n in pool if not is_unavailable(n, roster_map[n]["PGY"], "Nights", bi)
                    and not would_exceed_run(n, bi, "Nights")]
            if role == "Senior":
                pool = [n for n in pool if is_senior(n)]
            elif role == "Junior":
                pool = [n for n in pool if is_junior(n)]
            else:
                pool = [n for n in pool if is_intern(n)]
            if not relax_gap:
                pool = [n for n in pool if _nf_gap_ok(n, bi)]
            if not pool:
                return None
            if role == "Senior":
                keyf = lambda n: (Counter(schedule_df.loc[n]=="Nights")[True], repeat_penalty(n, bi, "Nights"),
                                  (_is_pgy5(n) if (avoid_pgy5_last and bi==12) else 0), n)
            else:
                keyf = lambda n: (Counter(schedule_df.loc[n]=="Nights")[True], repeat_penalty(n, bi, "Nights"), n)
            pool.sort(key=keyf)
            return pool[0]

        senior = _pick_night_role("Senior", relax_gap=False, avoid_pgy5_last=True)
        junior = _pick_night_role("Junior", relax_gap=False)
        intern = _pick_night_role("Intern", relax_gap=False)

        if senior is None:
            senior = _pick_night_role("Senior", relax_gap=True, avoid_pgy5_last=True)
        if junior is None:
            junior = _pick_night_role("Junior", relax_gap=True)
        if intern is None:
            intern = _pick_night_role("Intern", relax_gap=True)

        if senior and junior and intern:
            for role, pick in (("Senior", senior), ("Junior", junior), ("Intern", intern)):
                schedule_df.loc[pick, hdr] = "Nights"
                st.session_state["_nf_counts"][pick] += 1
                st.session_state["_last_nf"][pick] = bi
                st.session_state["_nf_role_counts"][pick][role] += 1
                log_reason(reasons, pick, hdr, f"Nights as {role} (role-fair; spacing respected; avoid consecutive)")

        # Floor (HARD): PGY-1 only - prefer interns with MORE Vascular so those with fewer are available for Vascular
        floor_interns = [n for n in names if roster_map[n]["PGY"]=="PGY-1"
                         and n not in locked_rows
                         and pd.isna(schedule_df.loc[n, hdr])
                         and not is_unavailable(n, roster_map[n]["PGY"], "Floor", bi)
                         and not would_exceed_run(n, bi, "Floor")]
        if floor_interns:
            # Prefer interns who already have MORE Vascular (negative count), so those with fewer stay available
            pick = sorted(floor_interns, key=lambda n:(-vasc_intern_counts[n],
                                                        int((schedule_df.loc[n]=="Floor").sum()),
                                                        repeat_penalty(n, bi, "Floor"),
                                                        rng.random(), n))[0]
            schedule_df.loc[pick, hdr] = "Floor"
            log_reason(reasons, pick, hdr, "Floor (PGY-1; prefer those with more Vascular)")

        # --------- Vascular baseline BEFORE RG/Gold seeding (attempt S/J/I; must include a Senior) ---------
        assign_first([n for n in names if is_senior(n) and n not in locked_rows], bi, "Vascular")  # Senior required (hard rule checked later)

        # Vascular intern: pick the intern with fewest Vascular assignments for even distribution
        avail_interns_for_vasc = [n for n in names if is_intern(n)
                                  and n not in locked_rows
                                  and pd.isna(schedule_df.loc[n, hdr])
                                  and not is_unavailable(n, roster_map[n]["PGY"], "Vascular", bi)
                                  and not would_exceed_run(n, bi, "Vascular")]
        if avail_interns_for_vasc:
            pick = sorted(avail_interns_for_vasc, key=lambda n:(vasc_intern_counts[n],
                                                                 int((schedule_df.loc[n]=="Vascular").sum()),
                                                                 repeat_penalty(n, bi, "Vascular"),
                                                                 rng.random(), n))[0]
            schedule_df.loc[pick, hdr] = "Vascular"
            vasc_intern_counts[pick] += 1
            log_reason(reasons, pick, hdr, "Vascular intern (PGY-1 fairness)")
        if count_team_here(bi,"Vascular") >= 2 and count_team_here(bi,"Vascular") < VASC_MAX:
            avail_for_vasc = [n for n in names if pd.isna(schedule_df.loc[n, hdr])
                              and n not in locked_rows
                              and is_junior(n)
                              and not is_unavailable(n, roster_map[n]["PGY"], "Vascular", bi)
                              and not would_violate_senior_cap(n, bi, "Vascular")
                              and not would_exceed_run(n, bi, "Vascular")]
            if avail_for_vasc:
                pick = sorted(avail_for_vasc, key=lambda n:(vasc_junior_counts[n],
                                                            int((schedule_df.loc[n]=="Vascular").sum()),
                                                            repeat_penalty(n, bi, "Vascular"),
                                                            n))[0]
                schedule_df.loc[pick, hdr] = "Vascular"
                vasc_junior_counts[pick] += 1
                log_reason(reasons, pick, hdr, "Vascular 3rd member (PGY-2/3 fairness)")

        # Breast (PGY-2+, cap=1)
        if count_team_here(bi, "Breast") < BREAST_MAX:
            assign_first([n for n in names if roster_map[n]["PGY"] in {"PGY-2","PGY-3","PGY-4","PGY-5"} and n not in locked_rows], bi, "Breast")

        # --------- Seed Gold & Red/Green to achieve S/J/I (attempts) and ensure Senior presence (hard) ---------
        def assign_with_fairness(pool, bi, label, role, counter_dict):
            """Assign using fairness counter as primary sort key."""
            hdr = headers[bi]
            cand = [n for n in pool if pd.isna(schedule_df.loc[n, hdr])
                    and not is_unavailable(n, roster_map[n]["PGY"], label, bi)
                    and not would_violate_senior_cap(n, bi, label)
                    and not would_exceed_run(n, bi, label)]
            if not cand:
                return None
            pick = sorted(cand, key=lambda n:(counter_dict[role][n],
                                               int((schedule_df.loc[n]==label).sum()),
                                               repeat_penalty(n, bi, label),
                                               senior_same_year_penalty(n, bi, label),
                                               rng.random(), n))[0]
            schedule_df.loc[pick, hdr] = label
            counter_dict[role][pick] += 1
            log_reason(reasons, pick, hdr, f"{label} {role} (fairness-tracked)")
            return pick

        # Gold: Senior, Junior, Intern (in that order) with fairness tracking
        assign_with_fairness([n for n in names if is_senior(n) and n not in locked_rows], bi, "Gold", "Senior", gold_counts)
        assign_with_fairness([n for n in names if is_junior(n) and n not in locked_rows], bi, "Gold", "Junior", gold_counts)
        assign_with_fairness([n for n in names if is_intern(n) and n not in locked_rows], bi, "Gold", "Intern", gold_counts)

        # Red/Green: Senior, Junior, Intern (in that order) with fairness tracking
        assign_with_fairness([n for n in names if is_senior(n) and n not in locked_rows], bi, "Red/Green", "Senior", rg_counts)
        assign_with_fairness([n for n in names if is_junior(n) and n not in locked_rows], bi, "Red/Green", "Junior", rg_counts)
        assign_with_fairness([n for n in names if is_intern(n) and n not in locked_rows], bi, "Red/Green", "Intern", rg_counts)

        # Priority fill with caps + hard min/max on Gold/RG; ICU min
        def fill_to(lbl,tgt,cap=None):
            while count_team_here(bi,lbl) < tgt and (cap is None or count_team_here(bi,lbl) < cap):
                avail = [x for x in names if pd.isna(schedule_df.loc[x, hdr])
                         and x not in locked_rows
                         and not is_unavailable(x, roster_map[x]["PGY"], lbl, bi)
                         and not would_violate_senior_cap(x, bi, lbl)
                         and not would_exceed_run(x, bi, lbl, limit=(3 if lbl=="Pittsburgh" else 2))]
                if lbl == "Breast":
                    avail = [x for x in avail if roster_map[x]["PGY"] in {"PGY-2","PGY-3","PGY-4","PGY-5"}]
                if not avail: break
                pick=sorted(avail, key=lambda n:(int((schedule_df.loc[n]==lbl).sum()),
                                                 repeat_penalty(n, bi, lbl),
                                                 senior_same_year_penalty(n, bi, lbl),
                                                 rng.random(), n))[0]
                schedule_df.loc[pick, hdr]=lbl
                log_reason(reasons, pick, hdr, f"Filled {lbl} to meet target/min/cap")

        # Hard minima
        fill_to("Gold", max(GOLD_MIN, svc_min.get("Gold",3)), cap=min(GOLD_MAX, GOLD_CAP_UI))
        fill_to("Red/Green", max(RG_MIN, svc_min.get("Red/Green",3)), cap=min(RG_MAX, RG_CAP_UI))
        fill_to("ICU", max(svc_min.get("ICU",1), 0), cap=1)

        # Fill to hard maximums (3–4)
        fill_to("Gold", GOLD_MAX, cap=min(GOLD_MAX, GOLD_CAP_UI))
        fill_to("Red/Green", RG_MAX, cap=min(RG_MAX, RG_CAP_UI))

        # HARD: everyone must have assignment — safety fill
        tries = 0
        while True:
            avail=[n for n in names if pd.isna(schedule_df.loc[n, hdr])]
            if not avail or tries >= 200: break
            n = avail.pop(0)
            # pick eligible destination prioritizing caps and avoiding consecutive & run limit
            choices = []
            for lbl, cap in (("Vascular", VASC_MAX),
                             ("Gold", GOLD_MAX),
                             ("Red/Green", RG_MAX),
                             ("Breast", 1),
                             ("Floor", None)):
                if is_unavailable(n, roster_map[n]["PGY"], lbl, bi): continue
                if lbl=="Floor" and roster_map[n]["PGY"]!="PGY-1": continue
                if cap is not None and count_team_here(bi, lbl) >= cap: continue
                if would_violate_senior_cap(n, bi, lbl): continue
                if would_exceed_run(n, bi, lbl, limit=(3 if lbl=="Pittsburgh" else 2)): continue
                if lbl=="Breast" and roster_map[n]["PGY"]=="PGY-1": continue
                choices.append(lbl)
            if not choices:
                if roster_map[n]["PGY"]!="PGY-1":
                    choices = [lbl for lbl in ["Gold","Red/Green","Vascular","Breast"] if (lbl!="Breast" or count_team_here(bi,"Breast")<1)]
                else:
                    choices = ["Floor","Vascular","Red/Green","Gold"]
            lbl = sorted(choices, key=lambda L:(repeat_penalty(n, bi, L), L))[0]
            schedule_df.loc[n, hdr] = lbl
            log_reason(reasons, n, hdr, f"Safety fill → {lbl} (hard: all residents assigned; minimal repeat)")
            tries += 1

        # Enforce Breast cap (rehoming extras if any)
        while count_team_here(bi,"Breast") > BREAST_MAX:
            extras = team(bi,"Breast")[BREAST_MAX:]  # keep first, move the rest
            moved = False
            for cand in extras:
                dest = None
                for lbl, cap in (("Gold", GOLD_MAX), ("Red/Green", RG_MAX), ("Vascular", VASC_MAX), ("Floor", None)):
                    if lbl=="Floor" and roster_map[cand]["PGY"]!="PGY-1": continue
                    if is_unavailable(cand, roster_map[cand]["PGY"], lbl, bi): continue
                    if cap is not None and count_team_here(bi,lbl) >= cap: continue
                    if would_violate_senior_cap(cand, bi, lbl): continue
                    if would_exceed_run(cand, bi, lbl): continue
                    dest = lbl; break
                if dest:
                    schedule_df.loc[cand, hdr] = dest
                    log_reason(reasons, cand, hdr, f"Rehomed from Breast to {dest} to enforce cap≤1")
                    moved = True
                    break
            if not moved:
                break  # cannot fix further without violating hard rules

        # Rebalance Gold vs Red/Green (keep diff ≤ 1)
        def rebalance_gold_rg():
            while abs(count_team_here(bi,"Red/Green") - count_team_here(bi,"Gold")) > 1:
                if count_team_here(bi,"Red/Green") > count_team_here(bi,"Gold") and count_team_here(bi,"Gold") < GOLD_MAX:
                    src = "Red/Green"; dst = "Gold"
                elif count_team_here(bi,"Gold") > count_team_here(bi,"Red/Green") and count_team_here(bi,"Red/Green") < RG_MAX:
                    src = "Gold"; dst = "Red/Green"
                else:
                    break
                cands = []
                for n in team(bi, src):
                    if is_unavailable(n, roster_map[n]["PGY"], dst, bi): continue
                    if would_violate_senior_cap(n, bi, dst): continue
                    if would_exceed_run(n, bi, dst): continue
                    cands.append(n)
                if not cands: break
                pick = sorted(cands, key=lambda n:(repeat_penalty(n, bi, dst),
                                                   int((schedule_df.loc[n]==dst).sum()),
                                                   senior_same_year_penalty(n, bi, dst),
                                                   n))[0]
                schedule_df.loc[pick, hdr] = dst
                log_reason(reasons, pick, hdr, f"Rebalanced {src}→{dst} to keep diff ≤ 1")
        rebalance_gold_rg()

        # (Attempt) Ensure Vascular has an intern via swap if possible (no longer a hard error)
        def vascular_has_intern() -> bool:
            return any(is_intern(n) for n in team(bi,"Vascular"))
        if not vascular_has_intern():
            vas = team(bi,"Vascular")
            jr_on_vas = [n for n in vas if is_junior(n)]
            if jr_on_vas:
                intern_pool = [n for n in names if is_intern(n) and schedule_df.loc[n, hdr] in {"Gold","Red/Green","Breast","Floor"}]
                if intern_pool:
                    give = intern_pool[0]
                    take = jr_on_vas[0]
                    old_lbl = schedule_df.loc[give, hdr]
                    if not exceeds_consecutive_limit(give, bi, "Vascular") and not exceeds_consecutive_limit(take, bi, old_lbl):
                        schedule_df.loc[give, hdr] = "Vascular"
                        schedule_df.loc[take, hdr] = old_lbl
                        log_reason(reasons, give, hdr, f"Swap to add Vascular intern (from {old_lbl})")
                        log_reason(reasons, take, hdr, f"Swap off Vascular to {old_lbl} (make room for intern)")

    # persist trackers
    st.session_state["assign_reasons"] = reasons
    return schedule_df

# --------------------------
# Optional Optimizer (ILP polish) using PuLP
# --------------------------
def polish_with_optimizer(schedule_df: pd.DataFrame, roster_df: pd.DataFrame, start_date: date, constraints: dict, time_limit_s: int = 8):
    if not HAS_PULP:
        st.info("Optimizer not available (PuLP not installed). Using greedy schedule.")
        return schedule_df

    names = list(schedule_df.index)
    headers = list(schedule_df.columns)
    roster = {r["Resident"]: r for _, r in normalize_roster_input(roster_df).iterrows()}

    # Hard targets
    GOLD_MIN, GOLD_MAX = 3, 4
    RG_MIN,   RG_MAX   = 3, 4
    VASC_MAX           = int(constraints.get("vascular_cap", 3))
    BREAST_MAX         = 1

    # penalties (kept for repeats etc.)
    w_nf_gap      = float(constraints.get("penalty_nf_gap", 3.0))
    w_pgy5_last   = float(constraints.get("penalty_pgy5_last", 5.0))
    w_repeat_svc  = float(constraints.get("penalty_consecutive_repeat", 2.0))

    SVC = ["Gold","Red/Green","Vascular","Breast","Chief","ICU","Floor","Nights","Elective","Pittsburgh","Vacation"]

    prob = pulp.LpProblem("Polish", pulp.LpMinimize)

    x = {}
    movable = {"Gold","Red/Green","Vascular","Breast","Floor"}
    for n in names:
        for bi, hdr in enumerate(headers):
            orig = str(schedule_df.loc[n, hdr])
            for s in SVC:
                if orig in {"Chief","ICU","Nights","Pittsburgh","Elective","Vacation"}:
                    val = 1 if s == orig else 0
                    x[(n,bi,s)] = pulp.LpVariable(f"x_{n}_{bi}_{s}", lowBound=val, upBound=val, cat="Binary")
                else:
                    if s in movable:
                        x[(n,bi,s)] = pulp.LpVariable(f"x_{n}_{bi}_{s}", lowBound=0, upBound=1, cat="Binary")
                    else:
                        val = 1 if s == orig else 0
                        x[(n,bi,s)] = pulp.LpVariable(f"x_{n}_{bi}_{s}", lowBound=val, upBound=val, cat="Binary")

    # Each (n,b) exactly one service
    for n in names:
        for bi, hdr in enumerate(headers):
            prob += pulp.lpSum(x[(n,bi,s)] for s in SVC) == 1

    # Team caps & balance within blocks
    for bi, hdr in enumerate(headers):
        gold_ct = pulp.lpSum(x[(n,bi,"Gold")] for n in names)
        rg_ct   = pulp.lpSum(x[(n,bi,"Red/Green")] for n in names)
        prob += gold_ct >= GOLD_MIN
        prob += gold_ct <= GOLD_MAX
        prob += rg_ct   >= RG_MIN
        prob += rg_ct   <= RG_MAX
        prob += pulp.lpSum(x[(n,bi,"Vascular")] for n in names) <= VASC_MAX
        prob += pulp.lpSum(x[(n,bi,"Breast")]   for n in names) <= BREAST_MAX

        # Senior presence (HARD): Gold, RG, Vascular must each have ≥1 Senior
        seniors = [n for n in names if role_of(n, roster)=="Senior"]
        prob += pulp.lpSum(x[(n,bi,"Gold")]      for n in seniors) >= 1
        prob += pulp.lpSum(x[(n,bi,"Red/Green")] for n in seniors) >= 1
        prob += pulp.lpSum(x[(n,bi,"Vascular")]  for n in seniors) >= 1

        # Senior cap ≤2 per service (across services)
        for svc in ["Gold","Red/Green","Vascular","Breast","ICU","Floor","Nights","Pittsburgh","Elective","Vacation"]:
            prob += pulp.lpSum(x[(n,bi,svc)] for n in seniors) <= 2

        # (No Pitts scheduling here; PGH is pinned by greedy)

    # Objective: soft penalties
    obj = 0

    # Penalty: PGY-5 on nights in last block
    bi_last = len(headers)-1
    pgy5s = [n for n in names if roster[n]["PGY"]=="PGY-5"]
    if pgy5s:
        obj += w_pgy5_last * pulp.lpSum(x[(n,bi_last,"Nights")] for n in pgy5s)

    # Penalty: consecutive repeats (excluding Pittsburgh which has allowed 3-run)
    y = {}
    for n in names:
        for bi in range(1, len(headers)):
            for s in SVC:
                y[(n,bi,s)] = pulp.LpVariable(f"y_rep_{n}_{bi}_{s}", lowBound=0, upBound=1, cat="Binary")
                prob += y[(n,bi,s)] <= x[(n,bi-1,s)]
                prob += y[(n,bi,s)] <= x[(n,bi,s)]
                prob += y[(n,bi,s)] >= x[(n,bi-1,s)] + x[(n,bi,s)] - 1
                if s != "Pittsburgh":
                    obj += w_repeat_svc * y[(n,bi,s)]

    prob += obj

    try:
        solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=max(1,int(time_limit_s)))
        prob.solve(solver)
    except Exception as e:
        st.warning(f"Optimizer error: {e}")
        return schedule_df

    out = schedule_df.copy()
    for n in names:
        for bi, hdr in enumerate(headers):
            for s in SVC:
                if pulp.value(x[(n,bi,s)]) >= 0.5:
                    out.loc[n, hdr] = s
                    break
    return out

# --------------------------
# Build Daily tabs
# --------------------------
def build_dailies_from_yearly(schedule_df: pd.DataFrame, start_date: date):
    blocks = build_blocks(start_date, 13)
    headers = [hdr_for_block(s,e) for (s,e) in blocks]
    names = list(schedule_df.index)
    dailies={}
    for bi,(s,e) in enumerate(blocks):
        cols=[]; d=s
        for _ in range(4):
            for wd in WEEKDAYS:
                cols.append((wd, d.strftime("%m/%d/%Y"))); d += timedelta(days=1)
        cols=pd.MultiIndex.from_tuples(cols, names=["Weekday","Date"])
        daily=pd.DataFrame(index=names, columns=cols, dtype=object)
        for n in names:
            base=schedule_df.loc[n, headers[bi]]
            daily.loc[n,:]=base
        wknd = daily.columns.get_level_values(0).isin(["Saturday","Sunday"])
        daily.loc[:, wknd] = ""  # weekends uncolored/empty except call text
        dailies[tab_name(s,e)] = daily
    return dailies

# --------------------------
# Vacation auto-fill from Roster
# --------------------------
def parse_vacation_spec(spec: str, blocks: list, start_date: date):
    """
    Parse a vacation specification and return a list of (block_index, weekday_dates) tuples.

    Supported formats:
    - "Block 3", "B3", "block3" - refers to block index 3 (1-based: Block 1 = first block)
    - "7/1/2025" or "07/01/2025" - a specific date (fills the week containing that date)
    - "7/1-7/5" or "7/1/2025-7/5/2025" - a date range

    Returns: list of (block_index, list_of_dates) where dates are strings like "07/01/2025"
    """
    if not isinstance(spec, str) or not spec.strip():
        return []

    spec = spec.strip()
    results = []

    # Try to parse as block reference: "Block 3", "B3", "block 3", etc.
    block_match = re.match(r"^(?:block\s*)?b?(\d+)$", spec.lower().strip())
    if block_match:
        block_num = int(block_match.group(1))
        # Convert from 1-based to 0-based index
        bi = block_num - 1
        if 0 <= bi < len(blocks):
            s, e = blocks[bi]
            # Get all weekdays in this block
            weekday_dates = []
            d = s
            while d <= e:
                if d.weekday() < 5:  # Monday=0 to Friday=4
                    weekday_dates.append(d.strftime("%m/%d/%Y"))
                d += timedelta(days=1)
            results.append((bi, weekday_dates))
        return results

    # Try to parse as date or date range
    date_patterns = [
        r"(\d{1,2}/\d{1,2}/\d{4})",  # 7/1/2025 or 07/01/2025
        r"(\d{1,2}/\d{1,2})",         # 7/1 (assumes current academic year)
    ]

    # Check for date range (date-date or date to date)
    range_match = re.match(r"(\d{1,2}/\d{1,2}(?:/\d{4})?)\s*[-–to]+\s*(\d{1,2}/\d{1,2}(?:/\d{4})?)", spec)
    if range_match:
        start_str, end_str = range_match.groups()
        try:
            # Parse start date
            if "/" in start_str and start_str.count("/") == 1:
                m, d = map(int, start_str.split("/"))
                y = start_date.year if m >= start_date.month else start_date.year + 1
                start_d = date(y, m, d)
            else:
                parts = start_str.split("/")
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                start_d = date(y, m, d)

            # Parse end date
            if "/" in end_str and end_str.count("/") == 1:
                m, d = map(int, end_str.split("/"))
                y = start_date.year if m >= start_date.month else start_date.year + 1
                end_d = date(y, m, d)
            else:
                parts = end_str.split("/")
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                end_d = date(y, m, d)

            # Find weekdays in range and which blocks they belong to
            d = start_d
            while d <= end_d:
                if d.weekday() < 5:  # Weekday
                    # Find which block this date belongs to
                    for bi, (bs, be) in enumerate(blocks):
                        if bs <= d <= be:
                            results.append((bi, [d.strftime("%m/%d/%Y")]))
                            break
                d += timedelta(days=1)
        except Exception:
            pass
        return results

    # Try single date
    single_date_match = re.match(r"(\d{1,2}/\d{1,2}(?:/\d{4})?)", spec)
    if single_date_match:
        date_str = single_date_match.group(1)
        try:
            if date_str.count("/") == 1:
                m, d = map(int, date_str.split("/"))
                y = start_date.year if m >= start_date.month else start_date.year + 1
                target_d = date(y, m, d)
            else:
                parts = date_str.split("/")
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                target_d = date(y, m, d)

            # Fill the week containing this date (Mon-Fri)
            # Find the Monday of that week
            monday = target_d - timedelta(days=target_d.weekday())
            weekday_dates = []
            for i in range(5):  # Mon-Fri
                wd = monday + timedelta(days=i)
                weekday_dates.append(wd.strftime("%m/%d/%Y"))

            # Find which block contains most of these dates
            for bi, (bs, be) in enumerate(blocks):
                if bs <= target_d <= be:
                    results.append((bi, weekday_dates))
                    break
        except Exception:
            pass

    return results

def apply_vacation_from_roster(dailies: dict, roster_df: pd.DataFrame, start_date: date):
    """
    Apply vacation entries from Roster's Vacation 1/2/3 columns to Daily Blocks.

    This fills in "Vacation" for the weekdays specified in each resident's vacation columns.
    """
    blocks = build_blocks(start_date, 13)
    headers = [hdr_for_block(s, e) for (s, e) in blocks]

    roster = normalize_roster_input(roster_df)
    roster = roster[roster["Include?"].str.upper().eq("Y")].reset_index(drop=True)

    for _, row in roster.iterrows():
        resident = row["Resident"]

        for vac_col in ["Vacation 1", "Vacation 2", "Vacation 3"]:
            if vac_col not in roster.columns:
                continue
            spec = row.get(vac_col, "")
            if not spec or str(spec).strip() == "" or str(spec).lower() == "nan":
                continue

            parsed = parse_vacation_spec(str(spec), blocks, start_date)
            for bi, date_list in parsed:
                if bi < 0 or bi >= len(blocks):
                    continue
                block_hdr = headers[bi]
                if block_hdr not in dailies:
                    continue
                daily_df = dailies[block_hdr]
                if resident not in daily_df.index:
                    continue

                # Fill in vacation for the specified dates
                for date_str in date_list:
                    for col in daily_df.columns:
                        # col is tuple (weekday, date_string)
                        if isinstance(col, tuple) and len(col) == 2:
                            weekday, col_date = col
                            if weekday in ("Saturday", "Sunday"):
                                continue  # Don't fill weekends
                            if col_date == date_str:
                                daily_df.loc[resident, col] = "Vacation"

    return dailies

# --------------------------
# Weekend call assignment
# --------------------------
def auto_assign_weekend_call(dailies: dict, schedule_df: pd.DataFrame, start_date: date, constraints: dict, roster_df: pd.DataFrame):
    r = normalize_roster_input(roster_df)
    r = r[r["Include?"].str.upper().eq("Y")].reset_index(drop=True)
    roster_map = {row["Resident"]: {"PGY": row["PGY"]} for _,row in r.iterrows()}
    names = [n for n in schedule_df.index if n in roster_map]
    def _role(n): return role_of(n, roster_map)

    blocks = build_blocks(start_date, 13)
    headers=[hdr_for_block(s,e) for (s,e) in blocks]
    excluded_base=set(map(norm_label, constraints.get("exclude_from_call", [])))

    call_year=Counter()
    # Track calls separately by role for fairer distribution within each group
    call_by_role = {"Senior": Counter(), "Junior": Counter(), "Intern": Counter()}
    # Track F/Su and Sa calls separately for even distribution of each type
    fsu_calls = Counter()  # F/Su calls per resident
    sa_calls = Counter()   # Sa calls per resident
    issues = st.session_state.setdefault("forced_call_issues", [])

    last_global_wknd = defaultdict(lambda: -999)
    def global_week_index(bi, wk): return bi*4 + wk

    # Random generator for fair tiebreaking
    call_rng = random.Random(int(constraints.get("random_seed", 0) or 0))

    def pick(pool, need, block_ctr, bi, wk, call_type_counter):
        pool = [n for n in pool if block_ctr[n]<2]

        # For Junior slot, strongly prefer actual juniors (PGY-2/3) for even distribution
        if need == "Junior":
            # First try to pick from actual juniors only, sorted by their call counts
            junior_pool = [n for n in pool if _role(n) == "Junior"]
            if junior_pool:
                # Primary sort: total calls (call_year) to balance across all juniors
                # Secondary: role-specific count, then call type, then spacing
                junior_pool.sort(key=lambda n:(call_year[n], call_by_role["Junior"][n], call_type_counter[n], -(global_week_index(bi,wk) - last_global_wknd[n]), call_rng.random(), n))
                return junior_pool[0]
            # Fall back to seniors if no juniors available
            senior_pool = [n for n in pool if _role(n) == "Senior"]
            if senior_pool:
                senior_pool.sort(key=lambda n:(call_year[n], call_type_counter[n], -(global_week_index(bi,wk) - last_global_wknd[n]), call_rng.random(), n))
                return senior_pool[0]
            return None

        # For Senior slot, pick from seniors sorted by their call counts
        if need == "Senior":
            senior_pool = [n for n in pool if _role(n) == "Senior"]
            if senior_pool:
                # Primary sort: total calls to balance across all seniors
                senior_pool.sort(key=lambda n:(call_year[n], call_by_role["Senior"][n], call_type_counter[n], -(global_week_index(bi,wk) - last_global_wknd[n]), call_rng.random(), n))
                return senior_pool[0]
            return None

        # For Intern slot, prefer actual interns
        if need == "Intern":
            intern_pool = [n for n in pool if _role(n) == "Intern"]
            if intern_pool:
                # Primary sort: total calls to balance across all interns
                intern_pool.sort(key=lambda n:(call_year[n], call_by_role["Intern"][n], call_type_counter[n], -(global_week_index(bi,wk) - last_global_wknd[n]), call_rng.random(), n))
                return intern_pool[0]
            # Fall back to juniors, then seniors
            for fallback_role in ["Junior", "Senior"]:
                fallback_pool = [n for n in pool if _role(n) == fallback_role]
                if fallback_pool:
                    fallback_pool.sort(key=lambda n:(call_year[n], call_type_counter[n], -(global_week_index(bi,wk) - last_global_wknd[n]), call_rng.random(), n))
                    return fallback_pool[0]
            return None

        return None

    for bi,(s,e) in enumerate(blocks):
        hdr=headers[bi]; daily=dailies[tab_name(s,e)]
        sat_cols=[c for c in daily.columns if c[0]=="Saturday"]
        sun_cols=[c for c in daily.columns if c[0]=="Sunday"]
        block_ctr=Counter()

        for wk in range(4):
            scol, sucol = sat_cols[wk], sun_cols[wk]
            base_pool=[n for n in names if norm_label(schedule_df.loc[n, hdr]) not in excluded_base]
            used=set()

            # F/Su - use fsu_calls counter for even F/Su distribution
            team=[]
            for need in ("Senior","Junior","Intern"):
                p=pick([n for n in base_pool if n not in used], need, block_ctr, bi, wk, fsu_calls)
                if p: team.append(p); used.add(p)
            if len(team)<3:
                pool_any=[n for n in names if n not in used and block_ctr[n]<2]
                while len(team)<3 and pool_any:
                    need=("Senior","Junior","Intern")[len(team)]
                    p=pick(pool_any, need, block_ctr, bi, wk, fsu_calls)
                    if not p: break
                    team.append(p); used.add(p); pool_any.remove(p)
                    issues.append(("Forced override", hdr, f"Week {wk+1}: {p} → F/Su to ensure coverage"))
            for n in team:
                daily.loc[n, [sucol]]="F/Su"; block_ctr[n]+=1; call_year[n]+=1; last_global_wknd[n]=global_week_index(bi,wk)
                call_by_role[_role(n)][n] += 1  # Track by role for fairness
                fsu_calls[n] += 1  # Track F/Su specifically
            if len(team)<3:
                issues.append(("Call coverage shortfall", hdr, f"Week {wk+1}: F/Su {len(team)}/3"))

            # Sa - use sa_calls counter for even Sa distribution
            team=[]
            for need in ("Senior","Junior","Intern"):
                p=pick([n for n in base_pool if n not in used], need, block_ctr, bi, wk, sa_calls)
                if p: team.append(p); used.add(p)
            if len(team)<3:
                pool_any=[n for n in names if n not in used and block_ctr[n]<2]
                while len(team)<3 and pool_any:
                    need=("Senior","Junior","Intern")[len(team)]
                    p=pick(pool_any, need, block_ctr, bi, wk, sa_calls)
                    if not p: break
                    team.append(p); used.add(p); pool_any.remove(p)
                    issues.append(("Forced override", hdr, f"Week {wk+1}: {p} → Sa to ensure coverage"))
            for n in team:
                daily.loc[n, [scol]]="Sa"; block_ctr[n]+=1; call_year[n]+=1; last_global_wknd[n]=global_week_index(bi,wk)
                call_by_role[_role(n)][n] += 1  # Track by role for fairness
                sa_calls[n] += 1  # Track Sa specifically
            if len(team)<3:
                issues.append(("Call coverage shortfall", hdr, f"Week {wk+1}: Sa {len(team)}/3"))

# --------------------------
# Effective yearly from dailies
# --------------------------
def effective_yearly_from_dailies(dailies: dict, start_date: date, index_order: list):
    blocks = build_blocks(start_date, 13)
    headers=[hdr_for_block(s,e) for (s,e) in blocks]
    eff=pd.DataFrame(index=index_order, columns=headers, dtype=object)
    for bi,(s,e) in enumerate(blocks):
        name=tab_name(s,e); df=dailies[name]
        wk_cols=[c for c in df.columns if c[0] in ["Monday","Tuesday","Wednesday","Thursday","Friday"]]
        for n in df.index:
            vals=[v for v in df.loc[n, wk_cols].values if isinstance(v,str) and v!=""]
            eff.loc[n, headers[bi]] = (pd.Series(vals).mode().iat[0] if vals else "")
    return eff

# --------------------------
# Checks with severity, codes, heatmap data
# --------------------------
def compute_checks(schedule_df: pd.DataFrame, dailies: dict, roster_df: pd.DataFrame, constraints: dict, start_date: date):
    if schedule_df is None or schedule_df.empty:
        empty = pd.DataFrame()
        return {"issues": empty, "summary": empty, "coverage": empty,
                "call": empty, "rot_counts": empty, "pgy_coverage": empty,
                "special_counts": empty, "svc_variance": empty}

    r = normalize_roster_input(roster_df)
    r = r[r["Include?"].str.upper().eq("Y")].reset_index(drop=True)
    order_keys={"PGY-5":0,"PGY-4":1,"PGY-3":2,"PGY-2":3,"PGY-1":4}
    r["ord"]=r["PGY"].map(order_keys); r=r.sort_values(["ord","Resident"]).drop(columns=["ord"]).reset_index(drop=True)
    roster_map={row["Resident"]: {"PGY": row["PGY"]} for _,row in r.iterrows()}

    names=[n for n in schedule_df.index if n in roster_map]
    headers=list(schedule_df.columns)
    blocks=build_blocks(start_date, 13)
    hdr_to_bi={hdr_for_block(s,e):i for i,(s,e) in enumerate(blocks)}
    svc_labels=["Gold","Red/Green","Vascular","Breast","Chief","ICU","Floor","Nights","Elective","Pittsburgh","Vacation"]

    issues=[]

    def add_issue(code, severity, hdr, msg):
        issues.append({"Code": code, "Severity": severity, "Block": hdr, "Details": msg})

    # Unassigned (ERROR)
    for hdr in headers:
        unassigned = [n for n in names if (pd.isna(schedule_df.loc[n, hdr]) or str(schedule_df.loc[n, hdr]).strip()=="")]
        for n in unassigned:
            add_issue("UNASSIGNED", "ERROR", hdr, f"{n} has no assignment")

    # Pittsburgh overlap & January (late) exclusion
    for idx,(s,e) in enumerate(blocks):
        hdr = headers[idx]
        pgh_count = int((schedule_df[hdr]=="Pittsburgh").sum())
        if pgh_count > 1:
            add_issue("PGH_OVERLAP", "ERROR", hdr, f"{pgh_count} residents in Pittsburgh (should be ≤ 1)")
        if block_has_late_january(s,e,start_date) and pgh_count>0:
            add_issue("PGH_JAN_LATE", "ERROR", hdr, "Pittsburgh scheduled in late January (Jan 16–31) not allowed")

    # PGY restriction on Pittsburgh (if enabled)
    if constraints.get("pg3_pittsburgh", True):
        for hdr in headers:
            bad = [n for n in names if schedule_df.loc[n, hdr]=="Pittsburgh" and roster_map[n]["PGY"]!="PGY-3"]
            for n in bad:
                add_issue("PGH_PGY_RULE", "ERROR", hdr, f"{n} ({roster_map[n]['PGY']}) on Pittsburgh (PGY-3 only)")

    # Pittsburgh run length must be exactly 3 (consecutive)
    for n in names:
        cur_len=0; start_idx=None
        for bi, hdr in enumerate(headers):
            if schedule_df.loc[n, hdr] == "Pittsburgh":
                if cur_len==0: start_idx=bi
                cur_len+=1
            else:
                if cur_len>0:
                    if cur_len != 3:
                        add_issue("PGH_RUN_LEN", "ERROR", headers[start_idx], f"{n} has Pittsburgh run of {cur_len} blocks (must be exactly 3)")
                    cur_len=0; start_idx=None
        if cur_len>0 and cur_len != 3:
            add_issue("PGH_RUN_LEN", "ERROR", headers[start_idx], f"{n} has Pittsburgh run of {cur_len} blocks (must be exactly 3)")

    # HARD: All PGY-3 must have exactly 3 Pittsburgh blocks (enforced/checked)
    for n in names:
        if roster_map[n]["PGY"] == "PGY-3":
            total_pgh = int((schedule_df=="Pittsburgh").loc[n].sum())
            if total_pgh != 3:
                add_issue("PGH_PGY3_REQUIRED", "ERROR", "(year)", f"{n} (PGY-3) has {total_pgh} Pittsburgh blocks (must be exactly 3)")

    # Per-block composition / caps / senior cap and team-size hard rules
    for hdr in headers:
        bi=hdr_to_bi.get(hdr,0)
        def svc(lbl): return [n for n in names if norm_label(schedule_df.loc[n, hdr])==norm_label(lbl)]
        gold      = svc("Gold")
        rg        = svc("Red/Green")
        breast    = svc("Breast")
        vas       = svc("Vascular")
        chief     = svc("Chief")
        icu       = svc("ICU")
        floor     = svc("Floor")
        nights    = svc("Nights")

        # Nights existence & composition (ERROR)
        roles = [role_of(n, roster_map) for n in nights]
        if len(nights) < 3:
            add_issue("NF_COVERAGE", "ERROR", hdr, f"Nights team size {len(nights)} (<3)")
        else:
            if Counter(roles) != Counter({"Senior":1,"Junior":1,"Intern":1}):
                add_issue("NF_COMPOSITION", "ERROR", hdr, f"Expected 1 Senior, 1 Junior, 1 Intern; got {Counter(roles)}")

        # Chief eligibility warnings (unchanged)
        if bi<6:
            for n in chief:
                if roster_map[n]["PGY"]=="PGY-2":
                    add_issue("CHIEF_H1_PGY2", "WARN", hdr, f"PGY-2 Chief: {n}")
        else:
            for n in chief:
                if roster_map[n]["PGY"]=="PGY-3":
                    add_issue("CHIEF_H2_PGY3", "WARN", hdr, f"PGY-3 Chief (fallback only): {n}")

        # ICU strict PGY requirements (ERROR)
        if bi<6:
            if not icu:
                add_issue("ICU_H1_NOCOVER", "ERROR", hdr, "ICU not covered by PGY-2 (required)")
            for n in icu:
                if roster_map[n]["PGY"]!="PGY-2":
                    add_issue("ICU_H1_BADPGY", "ERROR", hdr, f"{n} ({roster_map[n]['PGY']}) on ICU (PGY-2 only)")
        else:
            if not icu:
                add_issue("ICU_H2_NOCOVER", "ERROR", hdr, "ICU not covered by PGY-1 (required)")
            for n in icu:
                if roster_map[n]["PGY"]!="PGY-1":
                    add_issue("ICU_H2_BADPGY", "ERROR", hdr, f"{n} ({roster_map[n]['PGY']}) on ICU (PGY-1 only)")

        # Floor (ERROR if non-intern)
        for n in floor:
            if roster_map[n]["PGY"]!="PGY-1":
                add_issue("FLOOR_BADPGY", "ERROR", hdr, f"{n} ({roster_map[n]['PGY']}) on Floor (PGY-1 only)")

        # Senior-per-service cap (ERROR if >2)
        for svc_name, team_ in (("Gold",gold),("Red/Green",rg),("Vascular",vas),("Breast",breast),("ICU",icu),("Floor",floor),("Nights",nights),("Pittsburgh",svc("Pittsburgh")),("Elective",svc("Elective")),("Vacation",svc("Vacation"))):
            senior_ct = sum(1 for n in team_ if role_of(n, roster_map)=="Senior")
            if senior_ct > 2:
                add_issue("SR_CAP", "ERROR", hdr, f"{svc_name} has {senior_ct} seniors (>2)")

        # HARD: Gold & Red/Green size must be 3–4
        if not (3 <= len(gold) <= 4):
            add_issue("GOLD_SIZE", "ERROR", hdr, f"Gold has {len(gold)} (must be 3–4)")
        if not (3 <= len(rg) <= 4):
            add_issue("RG_SIZE", "ERROR", hdr, f"Red/Green has {len(rg)} (must be 3–4)")

        # HARD: Senior required on Gold, Red/Green, and Vascular
        if gold and not any(role_of(n, roster_map)=="Senior" for n in gold):
            add_issue("GOLD_NEEDS_SENIOR", "ERROR", hdr, "Gold has no Senior")
        if rg and not any(role_of(n, roster_map)=="Senior" for n in rg):
            add_issue("RG_NEEDS_SENIOR", "ERROR", hdr, "Red/Green has no Senior")
        if vas and not any(role_of(n, roster_map)=="Senior" for n in vas):
            add_issue("VASC_NEEDS_SENIOR", "ERROR", hdr, "Vascular has no Senior")

        # Attempt rule: Prefer S/J/I on Gold, RG, Vascular (WARN if missing Jr or Intern)
        def sji_warn(team_list, svcname):
            roles = {role_of(n, roster_map) for n in team_list}
            miss = [r for r in ["Senior","Junior","Intern"] if r not in roles]
            # Senior absence already flagged as ERROR; only warn for Junior/Intern attempts
            miss = [m for m in miss if m in {"Junior","Intern"}]
            if miss:
                add_issue("SJI_ATTEMPT", "WARN", hdr, f"{svcname} missing: {', '.join(miss)}")
        if gold: sji_warn(gold, "Gold")
        if rg:   sji_warn(rg, "Red/Green")
        if vas:  sji_warn(vas, "Vascular")

        # Avoid two seniors from same year on same team (WARN)
        def warn_same_year_seniors(team_list, svcname):
            seniors = [n for n in team_list if role_of(n, roster_map)=="Senior"]
            if len(seniors) >= 2:
                pgys = [roster_map[n]["PGY"] for n in seniors]
                if len(set(pgys)) == 1:
                    add_issue("SR_SAME_YEAR_PAIR", "WARN", hdr, f"{svcname} has two Seniors from same year ({pgys[0]})")
        warn_same_year_seniors(gold, "Gold")
        warn_same_year_seniors(rg, "Red/Green")
        warn_same_year_seniors(vas, "Vascular")

        # Balance warning if still imbalanced (diff > 1)
        if abs(len(rg) - len(gold)) > 1:
            add_issue("RG_GOLD_IMBALANCE", "WARN", hdr, f"Gold={len(gold)} vs Red/Green={len(rg)} (diff > 1)")

    # Night spacing (WARN)
    for n in names:
        last = None
        for bi, hdr in enumerate(headers):
            if schedule_df.loc[n, hdr] == "Nights":
                if last is not None and (bi - last) < 2:
                    add_issue("NF_SPACING", "WARN", hdr, f"{n} on NF < 1 block after {headers[last]}")
                last = bi

    # Consecutive >2 same rotation (ERROR) — exclude Pittsburgh (handled separately)
    for n in names:
        run = 1
        prev = None
        for bi, hdr in enumerate(headers):
            cur = schedule_df.loc[n, hdr]
            if cur == "Pittsburgh":
                run = 1; prev = None
                continue
            cur_norm = norm_label(cur)
            if cur_norm and prev == cur_norm:
                run += 1
            else:
                run = 1
            if run >= 3:
                add_issue("REPEAT_GT2", "ERROR", hdr, f"{n} has {run} consecutive {cur}")
            prev = cur_norm

    issues_df = pd.DataFrame(issues).sort_values(["Severity","Code","Block","Details"], ignore_index=True)
    summary_issue = (issues_df.value_counts(["Severity","Code"]).rename("Count").reset_index()
                     if not issues_df.empty else pd.DataFrame(columns=["Severity","Code","Count"]))

    # Coverage
    cov_rows=[]
    for hdr in headers:
        row={"Block": hdr}
        for lbl in svc_labels: row[lbl]=int((schedule_df[hdr]==lbl).sum())
        cov_rows.append(row)
    coverage_df=pd.DataFrame(cov_rows)

    # Call-load
    call_rows=[]
    for n in names:
        FSu=Sa=0
        for daily in dailies.values():
            vals = daily.loc[n].to_numpy()
            FSu += (vals=="F/Su").sum()
            Sa  += (vals=="Sa").sum()
        call_rows.append({"Resident": n, "PGY": roster_map[n]["PGY"], "F/Su": FSu, "Sa": Sa, "Total": FSu+Sa})
    order={"PGY-5":0,"PGY-4":1,"PGY-3":2,"PGY-2":3,"PGY-1":4}
    call_df=(pd.DataFrame(call_rows)
             .assign(_ord=lambda d:d["PGY"].map(order))
             .sort_values(["_ord","Total","Resident"], ascending=[True,False,True])
             .drop(columns=["_ord"]).reset_index(drop=True))

    # Rotation counts
    rot=[]
    for n in names:
        rrow={"Resident": n, "PGY": roster_map[n]["PGY"]}
        for lbl in svc_labels: rrow[lbl]=int((schedule_df==lbl).loc[n].sum())
        rot.append(rrow)
    rot_counts_df=(pd.DataFrame(rot)
                   .assign(_ord=lambda d:d["PGY"].map(order))
                   .sort_values(["_ord","Resident"])
                   .drop(columns=["_ord"]).reset_index(drop=True))

    # Per-PGY per block
    pgy_rows=[]
    for hdr in headers:
        for pgy in ["PGY-5","PGY-4","PGY-3","PGY-2","PGY-1"]:
            row={"Block":hdr,"PGY":pgy}
            for lbl in svc_labels:
                row[lbl]=sum(1 for n in names if roster_map[n]["PGY"]==pgy and schedule_df.loc[n,hdr]==lbl)
            pgy_rows.append(row)
    pgy_coverage_df=pd.DataFrame(pgy_rows)

    # Special counts
    spec=[]
    for n in names:
        spec.append({
            "Resident": n, "PGY": roster_map[n]["PGY"],
            "Night Float blocks": int((schedule_df=="Nights").loc[n].sum()),
            "Pittsburgh blocks": int((schedule_df=="Pittsburgh").loc[n].sum()),
            "Elective blocks": int((schedule_df=="Elective").loc[n].sum()),
        })
    special_counts_df=(pd.DataFrame(spec)
                       .assign(_ord=lambda d:d["PGY"].map(order))
                       .sort_values(["_ord","Resident"])
                       .drop(columns=["_ord"]).reset_index(drop=True))

    svc_totals={lbl:int((schedule_df==lbl).sum().sum()) for lbl in svc_labels}
    svc_variance_df=(pd.DataFrame([{"Service":k,"Total blocks":v} for k,v in svc_totals.items()])
                     .sort_values("Service").reset_index(drop=True))

    # Reasons & logs
    reasons = st.session_state.get("assign_reasons", [])
    reasons_df = pd.DataFrame(reasons) if reasons else pd.DataFrame(columns=["Resident","Block","Reason"])
    call_overrides = st.session_state.get("forced_call_issues", [])
    nf_overrides   = st.session_state.get("nf_overrides", [])
    out = {
        "issues": issues_df, "summary": summary_issue, "coverage": coverage_df,
        "call": call_df, "rot_counts": rot_counts_df, "pgy_coverage": pgy_coverage_df,
        "special_counts": special_counts_df, "svc_variance": svc_variance_df, "reasons": reasons_df
    }
    if call_overrides:
        out["call_overrides"] = pd.DataFrame(call_overrides, columns=["Type","Block","Details"])
    if nf_overrides:
        out["nf_overrides"] = pd.DataFrame(nf_overrides, columns=["Type","Block","Details"])

    out["heat_rotations"] = rot_counts_df
    out["heat_calls"] = call_df

    # Compliance checks: 4 days off per block, 3 vacation weeks, 2 personal days
    compliance_rows = []
    for n in names:
        # Count vacation weeks (from dailies - 5 consecutive weekdays with "Vacation")
        vacation_weeks = 0
        personal_days = 0
        days_off_by_block = {}

        for block_name, daily_df in dailies.items():
            if n not in daily_df.index:
                continue
            row_data = daily_df.loc[n]

            # Count days off in this block
            days_off = 0
            vacation_days_in_block = 0
            for (weekday, dstr), val in row_data.items():
                val_str = str(val).strip() if pd.notna(val) else ""
                # Weekend days off = blank (no F/Su or Sa call)
                if weekday in ("Saturday", "Sunday"):
                    if val_str == "" or val_str.lower() not in ("f/su", "sa"):
                        days_off += 1
                else:
                    # Weekdays off = blank or Vacation
                    if val_str == "" or val_str.lower() == "vacation":
                        days_off += 1
                    if val_str.lower() == "vacation":
                        vacation_days_in_block += 1
                # Count personal days anywhere
                if val_str.lower() == "personal day":
                    personal_days += 1
                    days_off += 1  # Personal days also count as days off

            days_off_by_block[block_name] = days_off
            # A vacation week is 5 consecutive vacation days (Mon-Fri)
            if vacation_days_in_block >= 5:
                vacation_weeks += vacation_days_in_block // 5

        # Check if all blocks have at least 4 days off
        blocks_with_4_days_off = sum(1 for d in days_off_by_block.values() if d >= 4)
        total_blocks = len(days_off_by_block)
        all_blocks_have_4_off = blocks_with_4_days_off == total_blocks if total_blocks > 0 else False

        compliance_rows.append({
            "Resident": n,
            "PGY": roster_map[n]["PGY"],
            "4+ Days Off (All Blocks)": "Yes" if all_blocks_have_4_off else f"No ({blocks_with_4_days_off}/{total_blocks})",
            "Vacation Weeks": vacation_weeks,
            "3 Vacation Weeks": "Yes" if vacation_weeks >= 3 else f"No ({vacation_weeks}/3)",
            "Personal Days": personal_days,
            "2 Personal Days": "Yes" if personal_days >= 2 else f"No ({personal_days}/2)",
        })

    compliance_df = (pd.DataFrame(compliance_rows)
                     .assign(_ord=lambda d: d["PGY"].map(order))
                     .sort_values(["_ord", "Resident"])
                     .drop(columns=["_ord"]).reset_index(drop=True))
    out["compliance"] = compliance_df

    return out

# --------------------------
# Amion CSV lint
# --------------------------
def lint_amion(dailies: dict):
    rows=[]
    for tab_name,df in dailies.items():
        for resident in df.index:
            for (weekday,dstr),val in df.loc[resident].items():
                if not isinstance(resident,str) or not resident.strip():
                    rows.append(("ERROR","Blank name", tab_name, dstr, resident, val))
                if not isinstance(dstr,str) or not re.match(r"^\d{2}/\d{2}/\d{4}$", dstr):
                    rows.append(("ERROR","Bad date", tab_name, dstr, resident, val))
                if isinstance(val,str) and ("," in val):
                    rows.append(("WARN","Comma in assignment", tab_name, dstr, resident, val))
    return pd.DataFrame(rows, columns=["Severity","Issue","BlockTab","Date","Resident","Assignment"])

# --------------------------
# Export helpers
# --------------------------
def _build_format_maps_xlsxwriter(workbook):
    """Create and cache XlsxWriter formats for each service color."""
    fmts = {}
    for key, hexcolor in PALETTE.items():
        fmts[key] = workbook.add_format({
            "bg_color": hexcolor,
            "font_color": "#FFFFFF" if key in WHITE_TEXT else "#000000",
            "bold": True if key in BOLD_TEXT else False,
        })
    return fmts

def _apply_coloring_xlsxwriter(ws, df: pd.DataFrame, start_row: int, start_col: int, fmts: dict, weekdays_only=False):
    """Write values back with formatting. start_row/col are 0-based in XlsxWriter."""
    nrows, ncols = df.shape
    for r in range(nrows):
        for c in range(ncols):
            val = df.iat[r, c]
            disp = "" if (pd.isna(val) or val == "") else str(val)
            fmt = None
            if weekdays_only and isinstance(df.columns, pd.MultiIndex):
                wd = df.columns[c][0]
                if wd in ("Saturday","Sunday"):
                    ws.write(start_row + r, start_col + c, disp)  # no color
                    continue
            key = norm_label(disp)
            if key in fmts:
                fmt = fmts[key]
            if fmt is not None:
                ws.write(start_row + r, start_col + c, disp, fmt)
            else:
                ws.write(start_row + r, start_col + c, disp)

def _apply_coloring_openpyxl(ws, df: pd.DataFrame, start_row: int, start_col: int, weekdays_only=False):
    """Fallback coloring with openpyxl. start_row/col are 1-based in openpyxl."""
    try:
        from openpyxl.styles import PatternFill, Font
    except Exception:
        return
    nrows, ncols = df.shape
    for r in range(nrows):
        for c in range(ncols):
            val = df.iat[r, c]
            disp = "" if (pd.isna(val) or val == "") else str(val)
            if weekdays_only and isinstance(df.columns, pd.MultiIndex):
                wd = df.columns[c][0]
                if wd in ("Saturday","Sunday"):
                    ws.cell(row=start_row + r, column=start_col + c, value=disp)
                    continue
            key = norm_label(disp)
            cell = ws.cell(row=start_row + r, column=start_col + c, value=disp)
            if key in PALETTE:
                hexcolor = PALETTE[key].lstrip("#")
                cell.fill = PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")
                font_color = "FFFFFFFF" if key in WHITE_TEXT else "FF000000"
                cell.font = Font(bold=(key in BOLD_TEXT), color=font_color)

def to_excel(schedule_df, dailies, checks) -> bytes:
    """Export with colors applied to Yearly and Daily sheets."""
    output=io.BytesIO()
    engine="xlsxwriter"
    try:
        import xlsxwriter as _xw  # noqa
    except Exception:
        engine="openpyxl"
    taken=set()
    with pd.ExcelWriter(output, engine=engine) as xw:
        # === Yearly (Schedule) ===
        schedule_name=safe_sheet_name("Schedule", taken)
        schedule_df.to_excel(xw, sheet_name=schedule_name)
        ws = xw.sheets[schedule_name]
        try:
            if engine == "xlsxwriter":
                ws.freeze_panes(1,1)
                ws.set_column(0, 0, 22)
                ws.set_column(1, schedule_df.shape[1], 14)
            else:
                ws.freeze_panes = "B2"
                ws.column_dimensions["A"].width = 22
        except Exception:
            pass

        if engine == "xlsxwriter":
            fmts = _build_format_maps_xlsxwriter(xw.book)
            _apply_coloring_xlsxwriter(ws, schedule_df, start_row=1, start_col=1, fmts=fmts, weekdays_only=False)
        else:
            _apply_coloring_openpyxl(ws, schedule_df, start_row=2, start_col=2, weekdays_only=False)

        # === Daily sheets ===
        for name,df in dailies.items():
            df2=df.loc[[i for i in df.index if isinstance(i,str) and i.strip().lower()!="none"]]
            sname=safe_sheet_name(name, taken)
            df2.to_excel(xw, sheet_name=sname)
            ws2 = xw.sheets[sname]
            try:
                if engine == "xlsxwriter":
                    ws2.freeze_panes(1,1)
                    ws2.set_column(0, 0, 22)
                    ws2.set_column(1, df2.shape[1], 12)
                else:
                    ws2.freeze_panes = "B2"
                    ws2.column_dimensions["A"].width = 22
            except Exception:
                pass

            if engine == "xlsxwriter":
                fmts = _build_format_maps_xlsxwriter(xw.book)
                header_depth = getattr(df2.columns, "nlevels", 1)
                start_row = header_depth
                start_col = 1
                _apply_coloring_xlsxwriter(ws2, df2, start_row=start_row, start_col=start_col, fmts=fmts, weekdays_only=True)
            else:
                header_depth = getattr(df2.columns, "nlevels", 1)
                start_row = 1 + header_depth
                start_col = 2
                _apply_coloring_openpyxl(ws2, df2, start_row=start_row, start_col=start_col, weekdays_only=True)

        # === Checks sheet ===
        checks_sheet=safe_sheet_name("Checks", taken)
        row=0
        for key in ["summary","issues","coverage","call","rot_counts","pgy_coverage","special_counts","svc_variance","reasons","call_overrides","nf_overrides"]:
            df=checks.get(key, pd.DataFrame())
            if df is not None and not df.empty:
                df.to_excel(xw, sheet_name=checks_sheet, index=False, startrow=row)
                try:
                    ws = xw.sheets[checks_sheet]
                    nrows = len(df)
                    if engine == "xlsxwriter" and "Severity" in df.columns:
                        col_idx = list(df.columns).index("Severity")
                        ws.conditional_format(row+1, col_idx, row+nrows, col_idx, {
                            "type": "text",
                            "criteria": "containing",
                            "value": "ERROR",
                            "format": xw.book.add_format({"bg_color":"#FFC7CE"})
                        })
                except Exception:
                    pass
                row += len(df) + 3
    return output.getvalue()

def to_amion_csv(dailies) -> bytes:
    rows=[]
    for tab_name,df in dailies.items():
        show=df.loc[[i for i in df.index if isinstance(i,str) and i.strip().lower()!="none"]]
        for resident in show.index:
            for (weekday,dstr),val in show.loc[resident].items():
                rows.append([resident,dstr,val])
    out=io.StringIO()
    pd.DataFrame(rows, columns=["Resident","Date","Assignment"]).to_csv(out, index=False)
    return out.getvalue().encode("utf-8")

# --------------------------
# ICS Calendar Generation
# --------------------------
def _ics_escape(text: str) -> str:
    """Escape special characters for ICS format."""
    if not text:
        return ""
    return text.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")

def _format_ics_date(d: date) -> str:
    """Format date as ICS DATE value (YYYYMMDD)."""
    return d.strftime("%Y%m%d")

def _generate_uid(resident: str, date_str: str, assignment: str) -> str:
    """Generate a unique ID for an event."""
    import hashlib
    content = f"{resident}-{date_str}-{assignment}"
    return hashlib.md5(content.encode()).hexdigest()[:16] + "@resident-scheduler"

def generate_ics_for_resident(resident: str, dailies: dict, schedule_df: pd.DataFrame = None) -> bytes:
    """
    Generate an ICS calendar file for a single resident.

    Creates all-day events for each assignment (rotation, call, vacation, etc.)
    """
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Resident Scheduler//Surgery Schedule//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{_ics_escape(resident)} - Surgery Schedule",
    ]

    # Track events to consolidate consecutive same-assignment days
    events = []  # list of (start_date, end_date, assignment, is_call)

    for block_name, df in dailies.items():
        if resident not in df.index:
            continue

        for (weekday, dstr), val in df.loc[resident].items():
            if not isinstance(val, str) or not val.strip():
                continue
            val = val.strip()
            if not val:
                continue

            # Parse date string (MM/DD/YYYY)
            try:
                parts = dstr.split("/")
                m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
                event_date = date(y, m, d)
            except Exception:
                continue

            # Determine if this is a call assignment
            is_call = val.lower() in ("f/su", "sa", "call")

            events.append((event_date, val, is_call))

    # Sort events by date
    events.sort(key=lambda x: x[0])

    # Consolidate consecutive days with same assignment (except calls)
    consolidated = []
    i = 0
    while i < len(events):
        start_date, assignment, is_call = events[i]

        # Calls are always single-day events
        if is_call:
            consolidated.append((start_date, start_date, assignment, True))
            i += 1
            continue

        # For rotations, find consecutive days with same assignment
        end_date = start_date
        j = i + 1
        while j < len(events):
            next_date, next_assign, next_is_call = events[j]
            # Check if next day and same assignment (not a call)
            if (next_date - end_date).days == 1 and next_assign == assignment and not next_is_call:
                end_date = next_date
                j += 1
            else:
                break

        consolidated.append((start_date, end_date, assignment, False))
        i = j

    # Generate ICS events
    now_stamp = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())

    for start_d, end_d, assignment, is_call in consolidated:
        uid = _generate_uid(resident, start_d.isoformat(), assignment)

        # For all-day events, DTEND should be the day AFTER the last day
        dtend = end_d + timedelta(days=1)

        # Create summary with call indicator
        if is_call:
            summary = f"📞 {assignment} Call"
            description = f"Weekend call assignment: {assignment}"
        else:
            summary = assignment
            description = f"Rotation: {assignment}"

        # Add color hint in description based on rotation
        color_hint = PALETTE.get(norm_label(assignment), "")
        if color_hint:
            description += f"\\nColor: {color_hint}"

        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now_stamp}",
            f"DTSTART;VALUE=DATE:{_format_ics_date(start_d)}",
            f"DTEND;VALUE=DATE:{_format_ics_date(dtend)}",
            f"SUMMARY:{_ics_escape(summary)}",
            f"DESCRIPTION:{_ics_escape(description)}",
        ])

        # Add categories for filtering
        if is_call:
            lines.append("CATEGORIES:Call,Weekend")
        else:
            lines.append(f"CATEGORIES:Rotation,{_ics_escape(assignment)}")

        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")

    # ICS requires CRLF line endings
    return "\r\n".join(lines).encode("utf-8")

def generate_all_ics_zip(dailies: dict, schedule_df: pd.DataFrame = None) -> bytes:
    """Generate a ZIP file containing ICS files for all residents."""
    import zipfile

    output = io.BytesIO()
    residents = set()
    for df in dailies.values():
        residents.update(df.index)
    residents = sorted([r for r in residents if isinstance(r, str) and r.strip() and r.lower() != "none"])

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
        for resident in residents:
            ics_data = generate_ics_for_resident(resident, dailies, schedule_df)
            # Sanitize filename
            safe_name = re.sub(r'[^\w\s-]', '', resident).strip().replace(' ', '_')
            zf.writestr(f"{safe_name}_schedule.ics", ics_data)

    return output.getvalue()

# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="Surgery Scheduler – Editable", layout="wide")
st.title("Surgery Scheduler – Editable Prototype")

# ---- Password gate (simple) ----
def _check_password():
    import streamlit as st
    if "APP_PASSWORD" not in st.secrets:
        st.error("Security is enabled but APP_PASSWORD is not set in this environment.")
        st.stop()
    if st.session_state.get("_authed", False):
        return True
    def _on_submit():
        expected = st.secrets["APP_PASSWORD"]
        ok = st.session_state.get("_pw_input", "") == expected
        st.session_state["_authed"] = bool(ok)
        if not ok:
            st.session_state["_pw_error"] = "😕 Incorrect password"
    with st.form("login", clear_on_submit=False):
        st.subheader("Login")
        st.text_input("Password", type="password", key="_pw_input")
        st.form_submit_button("Enter", on_click=_on_submit)
    if st.session_state.get("_pw_error"):
        st.error(st.session_state["_pw_error"])
    return st.session_state.get("_authed", False)

if not _check_password():
    st.stop()

with st.sidebar:
    if st.session_state.get("_authed") and st.button("Log out"):
        st.session_state["_authed"] = False
        st.experimental_rerun()
# ---- end password gate ----

if not HAS_ARROW:
    st.warning("Running in Lite mode (no pyarrow). Editable tables use CSV text areas; exports still work.")
# PuLP warning removed - optimizer feature is hidden from UI

with st.sidebar:
    st.markdown("### Academic Year")
    default_start=date(2026,6,29)
    ay_start=st.date_input("Start date (Block 1 begins Monday)", value=default_start)

    st.markdown("---"); st.markdown("### Constraints")

    st.markdown("**Team size caps**")
    st.number_input("Gold team cap", min_value=3, max_value=10, value=4, key="gold_cap")
    st.number_input("Red/Green team cap", min_value=3, max_value=10, value=4, key="rg_cap")
    st.number_input("Vascular team cap", min_value=2, max_value=5, value=3, key="vascular_cap")

    st.markdown("**Service minima (per block)**")
    gold_min = st.number_input("Gold min", 0, 6, 3, key="gold_min")
    rg_min   = st.number_input("Red/Green min", 0, 6, 3, key="rg_min")

    st.markdown("**Pittsburgh rotation**")
    st.text_input("Pittsburgh block blacklist (comma-separated indices 0–12)", value="", key="pgh_block_blacklist_txt")
    st.caption("Planner assigns **3 consecutive blocks** for **all PGY-3 residents** (capacity 1 per block), avoiding **late Jan (Jan 16–31)**.")

    st.markdown("**Weekend Call – Exclude these rotations**")
    st.multiselect(
        "Excluded from weekend call (may override as last resort)",
        ["Nights","Elective","Pittsburgh","Vacation","Breast","Vascular","Chief","ICU","Floor","Gold","Red/Green"],
        default=["Nights","Elective","Pittsburgh","Vacation"], key="exclude_from_call"
    )

    st.markdown("---")
    st.number_input("Target roster size", min_value=1, max_value=120, value=18, step=1, key="roster_size")
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("➕ Add blank row"):
            row = pd.DataFrame([{"Resident":"", "PGY":"PGY-1", "Prelim":"N", "Include?":"Y",
                                 "Notes":"","Vacation 1":"","Vacation 2":"","Vacation 3":""}])
            st.session_state.roster_table = pd.concat([st.session_state.get("roster_table", pd.DataFrame(columns=["Resident","PGY","Prelim","Include?","Notes","Vacation 1","Vacation 2","Vacation 3"])), row], ignore_index=True)
    with c2:
        if st.button("🧹 Remove empty/None rows"):
            if "roster_table" in st.session_state:
                rt = st.session_state.roster_table.copy()
                rt["Resident"] = rt["Resident"].astype(str)
                bad = rt["Resident"].str.strip().str.lower().isin(["", "none", "na", "n/a"])
                st.session_state.roster_table = rt[~bad].reset_index(drop=True)
    with c3:
        if st.button("↔ Resize to N"):
            if "roster_table" in st.session_state:
                st.session_state.roster_table = ensure_roster_size(st.session_state.roster_table, st.session_state.roster_size)
            else:
                st.session_state.roster_table = ensure_roster_size(pd.DataFrame(columns=["Resident","PGY","Prelim","Include?","Notes","Vacation 1","Vacation 2","Vacation 3"]), st.session_state.roster_size)

    # Hidden: these features are still functional but not exposed in UI
    # st.checkbox("Use optimizer (polish with ILP if available)", False, key="use_optimizer")
    # st.number_input("Optimizer time limit (sec)", 1, 30, 8, key="opt_time_limit")
    # st.number_input("Random seed (tie-breaks)", 0, 10_000, 0, key="random_seed")
    # scenario_name = st.text_input("Scenario name (save snapshot)", value="", key="scenario_name")
    # Set default values for hidden controls
    if "use_optimizer" not in st.session_state:
        st.session_state.use_optimizer = False
    if "opt_time_limit" not in st.session_state:
        st.session_state.opt_time_limit = 8
    if "random_seed" not in st.session_state:
        st.session_state.random_seed = 0

    # --------------------------
    # Schedule Versions
    # --------------------------
    st.markdown("---")
    st.markdown("### Schedule Versions")

    # Initialize snapshots storage
    if "schedule_snapshots" not in st.session_state:
        st.session_state.schedule_snapshots = {}

    # Save current schedule as snapshot
    snapshot_name = st.text_input("Version name:", key="snapshot_name_input", placeholder="e.g., Draft 1 - more ICU")
    if st.button("💾 Save Current as Version"):
        if snapshot_name.strip():
            save_schedule_snapshot(snapshot_name.strip())
            st.success(f"Saved: {snapshot_name}")
        else:
            st.warning("Enter a version name")

    # List saved versions
    if st.session_state.schedule_snapshots:
        st.markdown("**Saved Versions:**")
        for name, snap in st.session_state.schedule_snapshots.items():
            summary = get_snapshot_summary(snap)
            saved_time = time.strftime("%H:%M", time.localtime(summary["saved_at"])) if summary["saved_at"] else ""
            col1, col2 = st.columns([3, 1])
            with col1:
                st.caption(f"📋 {name} ({summary['residents']} residents) - {saved_time}")
            with col2:
                if st.button("Load", key=f"load_{name}"):
                    load_schedule_snapshot(name)
                    st.success(f"Loaded: {name}")
                    st.rerun()

    # Download/Upload versions
    st.markdown("---")
    st.markdown("**Import/Export**")

    # Download dropdown and button
    if st.session_state.schedule_snapshots:
        download_choice = st.selectbox("Download version:", list(st.session_state.schedule_snapshots.keys()), key="download_version_choice")
        json_data = export_snapshot_to_json(download_choice)
        if json_data:
            safe_name = re.sub(r'[^\w\s-]', '', download_choice).strip().replace(' ', '_')
            st.download_button(
                "⬇️ Download as File",
                data=json_data,
                file_name=f"schedule_{safe_name}.json",
                mime="application/json",
                key="download_snapshot_btn"
            )

    # Upload version
    uploaded_version = st.file_uploader("Upload version file:", type=["json"], key="upload_version_file")
    if uploaded_version:
        try:
            json_str = uploaded_version.read().decode("utf-8")
            imported = import_snapshot_from_json(json_str)
            if imported:
                import_name = imported["name"]
                # Avoid overwriting - add suffix if exists
                if import_name in st.session_state.schedule_snapshots:
                    import_name = f"{import_name} (imported)"
                st.session_state.schedule_snapshots[import_name] = imported
                st.success(f"Imported: {import_name}")
                st.rerun()
            else:
                st.error("Could not parse version file")
        except Exception as e:
            st.error(f"Error importing: {e}")

# Initialize default roster (display moved to Yearly tab)
_default_roster=pd.DataFrame([
    ["Avi Robinson","PGY-5","N","Y","","","",""],
    ["Kathleen Koesarie","PGY-5","N","Y","","","",""],
    ["Arruj Hassan","PGY-5","N","Y","","","",""],
    ["Shirin Siddiqi","PGY-5","N","Y","","","",""],
    ["Kayla Orr","PGY-4","N","Y","","","",""],
    ["Makayla Gologram","PGY-4","N","Y","","","",""],
    ["Adrianne Pellegrini","PGY-4","N","Y","","","",""],
    ["Zane Hamden","PGY-3","N","Y","","","",""],
    ["Lauren Delong","PGY-3","N","Y","","","",""],
    ["Brittany Steffens","PGY-3","N","Y","","","",""],
    ["Zoe Wecht","PGY-2","N","Y","","","",""],
    ["Jessica Marks","PGY-2","N","Y","","","",""],
    ["Jacob Allenabaugh","PGY-2","N","Y","","","",""],
    ["Intern 1","PGY-1","N","Y","","","",""],
    ["Intern 2","PGY-1","N","Y","","","",""],
    ["Intern 3","PGY-1","N","Y","","","",""],
    ["Intern 4 (Prelim)","PGY-1","Y","Y","","","",""],
    ["Intern 5 (Prelim)","PGY-1","Y","Y","","","",""],
], columns=["Resident","PGY","Prelim","Include?","Notes","Vacation 1","Vacation 2","Vacation 3"])

if "roster_table" not in st.session_state:
    st.session_state.roster_table = _default_roster.copy()

# Edit mode toggle (used across all tabs)
edit_mode = st.toggle("Edit mode", value=True, help="ON = tables editable; OFF = preview (weekday colors only)")

# Define generate_schedule function for use in Yearly tab
def generate_schedule_from_roster():
    """Generate schedule from current roster. Called from Yearly tab."""
    try:
        roster_raw = st.session_state.roster_table.copy()
        roster_raw["Resident"] = roster_raw["Resident"].astype(str)
        bad_names = roster_raw["Resident"].str.strip().str.lower().isin(["", "none", "na", "n/a"])
        roster_raw = roster_raw[~bad_names].reset_index(drop=True)
        roster_raw["PGY"] = roster_raw["PGY"].astype(str)
        roster_raw["PGY"] = roster_raw["PGY"].str.replace(r"\(.*?\)", "", regex=True)
        roster_raw["PGY"] = roster_raw["PGY"].str.upper().str.replace(r"PGY\s*([1-5])", r"PGY-\1", regex=True)
        roster_raw["PGY"] = roster_raw["PGY"].str.replace(r"^\s*([1-5])\s*$", r"PGY-\1", regex=True)
        roster_norm = normalize_roster_input(roster_raw)
        validate_roster(roster_norm)
    except Exception as e:
        st.error(f"Roster problem: {e}")
        return False

    # Parse Pittsburgh blacklist
    try:
        blk_txt = (st.session_state.pgh_block_blacklist_txt or "").strip()
        blk_list = []
        for tok in parse_csv_list(blk_txt):
            try:
                blk_list.append(int(tok))
            except Exception:
                pass
    except Exception:
        blk_list = []

    constraints = {
        "gold_cap": st.session_state.gold_cap,
        "rg_cap": st.session_state.rg_cap,
        "vascular_cap": st.session_state.vascular_cap,
        "service_minima": {"Gold": st.session_state.gold_min, "Red/Green": st.session_state.rg_min, "ICU": 1},
        "exclude_from_call": st.session_state.exclude_from_call,
        "enable_pittsburgh": True,
        "pg3_pittsburgh": True,
        "pgh_block_blacklist": blk_list,
        "eligibility_calendars": {},
        "random_seed": st.session_state.random_seed,
    }
    st.session_state.roster_df = roster_norm.copy()
    base_fixed = st.session_state.get("schedule_df", None)

    yearly = auto_generate_yearly(roster_norm.copy(), ay_start, constraints, base_fixed=base_fixed)
    if st.session_state.use_optimizer:
        yearly = polish_with_optimizer(yearly, roster_norm.copy(), ay_start, constraints, time_limit_s=st.session_state.opt_time_limit)

    dailies = build_dailies_from_yearly(yearly, ay_start)
    # Auto-fill vacation from Roster's Vacation 1/2/3 columns
    if st.session_state.get("auto_vacation", True):
        apply_vacation_from_roster(dailies, roster_norm, ay_start)
    if st.session_state.auto_call:
        auto_assign_weekend_call(dailies, yearly, ay_start, constraints, roster_norm)

    st.session_state.schedule_df = yearly
    st.session_state.dailies = dailies
    st.session_state.constraints = constraints
    st.session_state.schedule_df_effective = None
    save_schedule_to_cache(dailies, yearly, roster_norm)
    st.rerun()
    return True

# Utility: get a valid Yearly for checks/export
def get_effective_yearly_for_checks():
    sched = st.session_state.get("schedule_df", None)
    dailies = st.session_state.get("dailies", None)
    if sched is not None and not sched.empty:
        return sched
    if dailies:
        any_df = next(iter(dailies.values()))
        idx_order = list(any_df.index)
        return effective_yearly_from_dailies(dailies, ay_start, idx_order)
    return None

# Create tabs (always visible)
tabs = st.tabs(["Yearly (editable)","Daily Blocks (editable)","Checks","Export","Case Logs","Weekly Schedule"])

# Get schedule/dailies if they exist
schedule_df = st.session_state.get("schedule_df", None)
dailies = st.session_state.get("dailies", None)
constraints = st.session_state.get("constraints", {})
roster_df_ss = st.session_state.get("roster_df", st.session_state.roster_table)
has_schedule = schedule_df is not None and not schedule_df.empty

# Yearly tab
with tabs[0]:
    # Roster section (always visible in Yearly tab)
    with st.expander("Roster (editable)", expanded=not has_schedule):
        roster = show_table(st.session_state.roster_table, "roster_editor", editable=True, hide_index=True, index_name_hint="Resident")
        st.session_state.roster_table = roster.copy()

        st.checkbox("Auto-assign weekend call after edits", True, key="auto_call")
        st.checkbox("Auto-fill vacation dates from Roster columns", True, key="auto_vacation",
                   help="Fills 'Vacation' in Daily Blocks based on Vacation 1/2/3 columns. Accepts: 'Block 3', 'B3', '7/1/2025', or date ranges like '7/1-7/5'.")

        if st.button("Generate schedule from roster"):
            generate_schedule_from_roster()

    # Show schedule if it exists
    if has_schedule:
        st.markdown("#### Yearly Schedule")
        target = st.selectbox("Jump to resident", ["(all)"] + list(schedule_df.index), index=0, key="jump_resident")
        view_df = schedule_df if target == "(all)" else schedule_df.loc[[target]]
        if edit_mode:
            yearly_edit = view_df.reset_index().rename(columns={"index":"Resident"})
            # Get all block columns (everything except "Resident") for dropdown menus
            block_columns = [c for c in yearly_edit.columns if c != "Resident"]
            edited = show_table(yearly_edit, "yearly_editor", editable=True, hide_index=True,
                              index_name_hint="Resident", dropdown_columns=block_columns,
                              dropdown_options=ROTATION_OPTIONS)

            col1, col2 = st.columns(2)
            with col1:
                if st.button("Apply Yearly edits → rebuild Daily / Checks / Export"):
                    new_yearly = edited.set_index("Resident") if "Resident" in edited.columns else edited.copy()
                    if target != "(all)":
                        full = schedule_df.copy()
                        full.loc[new_yearly.index, :] = new_yearly.values
                        new_yearly = full
                    st.session_state.schedule_df = new_yearly
                    new_dailies = build_dailies_from_yearly(new_yearly, ay_start)
                    # Auto-fill vacation from Roster's Vacation 1/2/3 columns
                    if st.session_state.get("auto_vacation", True):
                        apply_vacation_from_roster(new_dailies, roster_df_ss, ay_start)
                    if st.session_state.auto_call:
                        auto_assign_weekend_call(new_dailies, new_yearly, ay_start, constraints, roster_df_ss)
                    st.session_state.dailies = new_dailies
                    st.session_state.schedule_df_effective = None
                    save_schedule_to_cache(new_dailies, new_yearly, roster_df_ss)
                    st.success("Applied Yearly edits.")
                    st.rerun()
            with col2:
                if st.button("Re-balance schedule (keep Pittsburgh/Elective, adjust others)"):
                    new_yearly = edited.set_index("Resident") if "Resident" in edited.columns else edited.copy()
                    if target != "(all)":
                        full = schedule_df.copy()
                        full.loc[new_yearly.index, :] = new_yearly.values
                        new_yearly = full

                    # Create base_fixed with only Pittsburgh, Elective, and Vacation (manually entered rotations to keep)
                    # Clear other cells so the scheduler can re-balance them
                    base_fixed = new_yearly.copy()
                    keep_rotations = {"Pittsburgh", "Elective", "Vacation"}
                    for col in base_fixed.columns:
                        for idx in base_fixed.index:
                            val = str(base_fixed.loc[idx, col]).strip()
                            if val not in keep_rotations:
                                base_fixed.loc[idx, col] = ""

                    # Re-run scheduler with base_fixed to fill in other rotations fairly
                    rebalanced = auto_generate_yearly(roster_df_ss.copy(), ay_start, constraints, base_fixed=base_fixed)
                    new_dailies = build_dailies_from_yearly(rebalanced, ay_start)
                    # Auto-fill vacation from Roster's Vacation 1/2/3 columns
                    if st.session_state.get("auto_vacation", True):
                        apply_vacation_from_roster(new_dailies, roster_df_ss, ay_start)
                    if st.session_state.auto_call:
                        auto_assign_weekend_call(new_dailies, rebalanced, ay_start, constraints, roster_df_ss)

                    st.session_state.schedule_df = rebalanced
                    st.session_state.dailies = new_dailies
                    st.session_state.schedule_df_effective = None
                    save_schedule_to_cache(new_dailies, rebalanced, roster_df_ss)
                    st.success("Re-balanced schedule (kept Pittsburgh/Elective/Vacation, adjusted others).")
                    st.rerun()
        else:
            vis = view_df.reset_index().rename(columns={"index":"Resident"})
            cols = vis.columns[1:]
            st.dataframe(vis.style.applymap(style_each_cell, subset=cols),
                         use_container_width=True, hide_index=True)
    else:
        st.info("No schedule yet. Edit the roster above and click 'Generate schedule from roster' to create one.")

    # Version Comparison Section
    if st.session_state.get("schedule_snapshots"):
        with st.expander("📊 Compare Versions", expanded=False):
            snapshot_names = list(st.session_state.schedule_snapshots.keys())

            if len(snapshot_names) >= 1:
                col1, col2 = st.columns(2)
                with col1:
                    compare_left = st.selectbox("Version A:", ["(Current)"] + snapshot_names, key="compare_left")
                with col2:
                    compare_right = st.selectbox("Version B:", ["(Current)"] + snapshot_names, index=min(1, len(snapshot_names)), key="compare_right")

                # Get the two schedules to compare
                def get_schedule_for_compare(choice):
                    if choice == "(Current)":
                        return st.session_state.get("sched_df", pd.DataFrame())
                    else:
                        snap = st.session_state.schedule_snapshots.get(choice, {})
                        return snap.get("schedule_df", pd.DataFrame())

                left_df = get_schedule_for_compare(compare_left)
                right_df = get_schedule_for_compare(compare_right)

                if not left_df.empty and not right_df.empty:
                    st.markdown(f"**Differences between {compare_left} and {compare_right}:**")

                    # Find differences
                    differences = []
                    common_residents = set(left_df.index) & set(right_df.index)
                    common_blocks = [c for c in left_df.columns if c in right_df.columns]

                    for resident in sorted(common_residents):
                        for block in common_blocks:
                            left_val = str(left_df.loc[resident, block]) if resident in left_df.index else ""
                            right_val = str(right_df.loc[resident, block]) if resident in right_df.index else ""
                            if left_val != right_val:
                                differences.append({
                                    "Resident": resident,
                                    "Block": block,
                                    compare_left: left_val,
                                    compare_right: right_val
                                })

                    if differences:
                        diff_df = pd.DataFrame(differences)
                        st.dataframe(diff_df, use_container_width=True, hide_index=True)
                        st.caption(f"{len(differences)} difference(s) found")
                    else:
                        st.success("No differences found - schedules are identical!")

                    # Summary stats comparison
                    st.markdown("**Summary Comparison:**")
                    def count_rotations(df):
                        counts = {}
                        for col in df.columns:
                            for val in df[col].dropna():
                                val_str = str(val).strip()
                                if val_str:
                                    counts[val_str] = counts.get(val_str, 0) + 1
                        return counts

                    left_counts = count_rotations(left_df)
                    right_counts = count_rotations(right_df)
                    all_rotations = sorted(set(left_counts.keys()) | set(right_counts.keys()))

                    summary_data = []
                    for rot in all_rotations:
                        left_c = left_counts.get(rot, 0)
                        right_c = right_counts.get(rot, 0)
                        diff = right_c - left_c
                        diff_str = f"+{diff}" if diff > 0 else str(diff) if diff < 0 else "="
                        summary_data.append({
                            "Rotation": rot,
                            compare_left: left_c,
                            compare_right: right_c,
                            "Diff": diff_str
                        })

                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(summary_df, use_container_width=True, hide_index=True)
                else:
                    st.info("Select two versions with schedule data to compare")
            else:
                st.info("Save at least one version to compare (use sidebar)")

# Daily tab
with tabs[1]:
    if dailies:
        st.markdown("#### Daily block tabs (Weekends uncolored; Sa / F/Su text only)")
        new_dailies={}
        for name,df in dailies.items():
            st.markdown(f"**{name}**")
            if edit_mode:
                # Get all columns for dropdown menus (all daily columns)
                daily_columns = list(df.columns)
                edited = show_table(df, f"edit_{name}", editable=True, daily=True, index_name_hint="Resident",
                                   dropdown_columns=daily_columns, dropdown_options=DAILY_OPTIONS)
                new_dailies[name]=edited.copy()
            else:
                weekday_cols=[c for c in df.columns if c[0] in ["Monday","Tuesday","Wednesday","Thursday","Friday"]]
                st.dataframe(df.style.applymap(style_each_cell, subset=weekday_cols),
                             use_container_width=True)
        if edit_mode and st.button("Apply Daily edits → recompute Checks & Export"):
            st.session_state.dailies = new_dailies
            eff_yearly = effective_yearly_from_dailies(new_dailies, ay_start, list(schedule_df.index))
            st.session_state.schedule_df_effective = eff_yearly
            save_schedule_to_cache(new_dailies, eff_yearly, roster_df_ss)
            st.success("Applied Daily edits and derived effective Yearly.")
    else:
        st.info("No schedule available yet. Generate a schedule in the Yearly tab first.")

# Checks tab
with tabs[2]:
    st.markdown("#### Checks & Balance Tables")
    base_yearly = st.session_state.get("schedule_df_effective", None) or get_effective_yearly_for_checks()
    if base_yearly is None or base_yearly.empty:
        st.info("No schedule available yet. Generate a schedule first.")
    else:
        checks = compute_checks(base_yearly, st.session_state.get("dailies", {}), roster_df_ss, constraints, ay_start)
        c1,c2,c3 = st.columns([1.1,1.25,1.25])
        with c1:
            st.markdown("**Summary (Severity × Code)**")
            st.dataframe(checks["summary"], use_container_width=True, hide_index=True)
        with c2:
            st.markdown("**Per-block coverage**")
            st.dataframe(checks["coverage"], use_container_width=True, hide_index=True)
        with c3:
            st.markdown("**Call-load by resident**")
            st.dataframe(checks["call"], use_container_width=True, hide_index=True)

        # Compliance checks (4 days off, vacation weeks, personal days)
        st.markdown("**Compliance Checks (Days Off / Vacation / Personal Days)**")
        if "compliance" in checks and not checks["compliance"].empty:
            st.dataframe(checks["compliance"], use_container_width=True, hide_index=True)
        else:
            st.info("No compliance data available.")

        st.markdown("**Issues (detailed)**"); st.dataframe(checks["issues"], use_container_width=True, hide_index=True)
        st.markdown("**Per-resident rotation counts (PGY-sorted)**"); st.dataframe(checks["rot_counts"], use_container_width=True, hide_index=True)
        st.markdown("**Per-PGY coverage per block**"); st.dataframe(checks["pgy_coverage"], use_container_width=True, hide_index=True)
        st.markdown("**Special blocks per resident**"); st.dataframe(checks["special_counts"], use_container_width=True, hide_index=True)
        st.markdown("**Service totals across the year**"); st.dataframe(checks["svc_variance"], use_container_width=True, hide_index=True)
        if "reasons" in checks and not checks["reasons"].empty:
            st.markdown("**Why these assignments?**"); st.dataframe(checks["reasons"], use_container_width=True, hide_index=True)
        if "call_overrides" in checks:
            st.markdown("**Weekend call overrides (forced for coverage)**")
            st.dataframe(checks["call_overrides"], use_container_width=True, hide_index=True)
        if "nf_overrides" in checks:
            st.markdown("**Night Float overrides & warnings**")
            st.dataframe(checks["nf_overrides"], use_container_width=True, hide_index=True)

        # Fairness Charts
        st.markdown("---")
        st.markdown("#### Fairness Charts")
        st.caption("All charts sorted by PGY level (PGY-5 first, then PGY-4, PGY-3, PGY-2, PGY-1).")

        # PGY sort order for all charts
        pgy_order = {"PGY-5": 0, "PGY-4": 1, "PGY-3": 2, "PGY-2": 3, "PGY-1": 4}

        # Helper to get sorted resident order
        def get_resident_order(df):
            df = df.copy()
            df["_pgy_ord"] = df["PGY"].map(pgy_order)
            df_sorted = df.sort_values(["_pgy_ord", "Resident"])
            return df_sorted["Resident"].tolist()

        # Weekend Call Load Chart (by type)
        st.markdown("**Weekend Call Load by Resident (F/Su vs Sa)**")
        call_data = checks["call"].copy()
        resident_order = get_resident_order(call_data)
        call_melted = call_data.melt(id_vars=["Resident", "PGY"], value_vars=["F/Su", "Sa"], var_name="Call Type", value_name="Count")
        call_chart = alt.Chart(call_melted).mark_bar().encode(
            x=alt.X("Resident:N", sort=resident_order, title="Resident"),
            y=alt.Y("Count:Q", title="Count"),
            color=alt.Color("Call Type:N"),
            xOffset="Call Type:N"
        ).properties(height=300)
        st.altair_chart(call_chart, use_container_width=True)

        # Weekend Call Load Total Chart
        st.markdown("**Weekend Call Load Total by Resident**")
        call_total_chart = alt.Chart(call_data).mark_bar().encode(
            x=alt.X("Resident:N", sort=resident_order, title="Resident"),
            y=alt.Y("Total:Q", title="Total Calls")
        ).properties(height=250)
        st.altair_chart(call_total_chart, use_container_width=True)

        # Special Blocks Chart
        st.markdown("**Special Blocks by Resident**")
        special_data = checks["special_counts"].copy()
        resident_order = get_resident_order(special_data)
        special_melted = special_data.melt(id_vars=["Resident", "PGY"], value_vars=["Night Float blocks", "Pittsburgh blocks", "Elective blocks"], var_name="Block Type", value_name="Count")
        special_chart = alt.Chart(special_melted).mark_bar().encode(
            x=alt.X("Resident:N", sort=resident_order, title="Resident"),
            y=alt.Y("Count:Q", title="Count"),
            color=alt.Color("Block Type:N"),
            xOffset="Block Type:N"
        ).properties(height=300)
        st.altair_chart(special_chart, use_container_width=True)

        # Rotation Distribution Chart
        st.markdown("**Rotation Distribution by Resident (All Rotations)**")
        rot_data = checks["rot_counts"].copy()
        rot_cols = [c for c in rot_data.columns if c not in ["Resident", "PGY"]]
        resident_order = get_resident_order(rot_data)
        rot_melted = rot_data.melt(id_vars=["Resident", "PGY"], value_vars=rot_cols, var_name="Rotation", value_name="Count")
        rot_chart = alt.Chart(rot_melted).mark_bar().encode(
            x=alt.X("Resident:N", sort=resident_order, title="Resident"),
            y=alt.Y("Count:Q", title="Count"),
            color=alt.Color("Rotation:N"),
            xOffset="Rotation:N"
        ).properties(height=400)
        st.altair_chart(rot_chart, use_container_width=True)

        # Individual Rotation Charts
        st.markdown("---")
        st.markdown("#### Individual Rotation Comparisons")
        st.caption("Each chart shows how many blocks each resident has for that rotation (sorted by PGY).")

        # Create individual bar chart for each rotation (sorted by PGY)
        for rotation in rot_cols:
            st.markdown(f"**{rotation}**")
            single_chart = alt.Chart(rot_data).mark_bar().encode(
                x=alt.X("Resident:N", sort=resident_order, title="Resident"),
                y=alt.Y(f"{rotation}:Q", title="Blocks")
            ).properties(height=250)
            st.altair_chart(single_chart, use_container_width=True)

# Export tab
with tabs[3]:
    base_yearly = st.session_state.get("schedule_df_effective", None) or get_effective_yearly_for_checks()
    if base_yearly is None or base_yearly.empty:
        st.info("No schedule available to export yet.")
    else:
        lint = lint_amion(st.session_state.get("dailies", {}))
        if not lint.empty:
            st.markdown("**Amion CSV Lint**"); st.dataframe(lint, use_container_width=True, hide_index=True)

        checks_now = compute_checks(base_yearly, st.session_state.get("dailies", {}), roster_df_ss, constraints, ay_start)

        # Spreadsheet exports
        st.markdown("#### Spreadsheet Exports")
        col_xlsx, col_csv = st.columns(2)
        with col_xlsx:
            xlsx_bytes = to_excel(base_yearly, st.session_state.get("dailies", {}), checks_now)
            st.download_button("⬇️ Download Excel (.xlsx)", data=xlsx_bytes,
                               file_name="ResidentSchedule_26-27_Editable.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col_csv:
            csv_bytes = to_amion_csv(st.session_state.get("dailies", {}))
            st.download_button("⬇️ Download Amion-style CSV (.csv)", data=csv_bytes,
                               file_name="ResidentSchedule_26-27_Amion.csv", mime="text/csv")

        # Calendar Subscriptions Section
        st.markdown("---")
        st.markdown("#### 📅 Calendar Subscriptions (Auto-Updating)")
        st.caption("Residents subscribe to a URL - their calendars update automatically when you change the schedule.")

        export_dailies = st.session_state.get("dailies", {})
        if export_dailies:
            # Get list of residents
            export_residents = set()
            for df in export_dailies.values():
                export_residents.update(df.index)
            export_residents = sorted([r for r in export_residents if isinstance(r, str) and r.strip() and r.lower() != "none"])

            st.markdown("**Subscription URLs for each resident:**")

            # Generate subscription URLs table
            sub_data = []
            for resident in export_residents:
                safe_name = re.sub(r'[^\w\s-]', '', resident).strip().replace(' ', '_')
                sub_url = f"/Calendar_Feed?resident={safe_name}"
                sub_data.append({"Resident": resident, "Subscription URL": sub_url})

            sub_df = pd.DataFrame(sub_data)
            st.dataframe(sub_df, use_container_width=True, hide_index=True)

            st.page_link("pages/1_📅_Calendar_Feed.py", label="📅 Open Calendar Feed Page", icon="🔗")

            with st.expander("How residents subscribe to calendar updates"):
                st.markdown("""
**For Residents - How to Subscribe:**

1. **Get your URL** from the table above (ask your admin for the full app URL)
2. **Full URL format:** `https://your-app.streamlit.app/Calendar_Feed?resident=Your_Name`

**Google Calendar:**
1. Open [calendar.google.com](https://calendar.google.com)
2. Click **+** next to "Other calendars" → **From URL**
3. Paste your subscription URL
4. Click **Add calendar**
5. ✅ Your calendar will auto-update every 12-24 hours

**Microsoft Outlook:**
1. Go to Calendar → **Add calendar** → **Subscribe from web**
2. Paste your subscription URL
3. Click **Import**

**Apple Calendar:**
1. File → **New Calendar Subscription**
2. Paste your subscription URL
3. Set refresh frequency (Auto, Every day, etc.)
4. Click **Subscribe**

**Note:** Updates aren't instant - calendar apps refresh every few hours to few days depending on the app.
                """)

# Case Logs tab
with tabs[4]:
    st.markdown("## 📊 ACGME Case Logs Dashboard")
    st.caption("Upload ACGME case log exports (PDF or Excel) to track resident progress.")

    # Initialize case log storage from cache
    # case_logs stores summary data: {resident: {category: {"total": N, "minimum": M, "subcategories": {...}}}}
    if "case_logs" not in st.session_state:
        st.session_state.case_logs = load_case_logs_from_cache()

    def is_summary_format(df):
        """Check if the DataFrame is an ACGME summary report (categories with totals)."""
        if df is None or df.empty:
            return False
        cols_lower = [str(c).lower() for c in df.columns]
        # Summary format typically has: Category, Minimum, and resident name columns
        has_category = any('category' in c or 'defined' in c for c in cols_lower)
        has_minimum = any('minimum' in c or 'min' in c for c in cols_lower)
        # Check if first column looks like category names (not dates or CPT codes)
        first_col_vals = df.iloc[:, 0].dropna().astype(str).head(5)
        has_text_categories = any(len(v) > 5 and not v.replace('.', '').isdigit() for v in first_col_vals)
        return (has_category or has_minimum) and has_text_categories

    def import_summary_file(df, filename):
        """Import an ACGME summary report file."""
        parsed = parse_acgme_summary_report(df)
        imported_residents = []
        for resident, categories in parsed.items():
            if resident and categories:
                # Clean resident name
                res_name = str(resident).strip()
                if res_name.lower() not in ['minimum', 'min', 'category', 'nan', '']:
                    st.session_state.case_logs[res_name] = categories
                    imported_residents.append(res_name)
        return imported_residents

    def process_imported_files(new_imports, source_name):
        """Process imported files from any source."""
        total_imported = 0
        for filename, df in new_imports:
            if is_summary_format(df):
                imported = import_summary_file(df, filename)
                if imported:
                    st.toast(f"📥 {source_name}: Imported {', '.join(imported)}")
                    total_imported += len(imported)
            else:
                # Try to parse anyway
                df.columns = df.columns.str.strip()
                resident_col = None
                for c in df.columns:
                    if any(x in c.lower() for x in ['resident', 'name', 'trainee', 'fellow']):
                        resident_col = c
                        break
                if resident_col is None and len(df.columns) > 0:
                    resident_col = df.columns[0]
                if resident_col:
                    for resident in df[resident_col].unique():
                        if pd.isna(resident) or str(resident).strip() == "":
                            continue
                        res_name = str(resident).strip()
                        resident_df = df[df[resident_col] == resident].copy()
                        if res_name not in st.session_state.case_logs:
                            st.session_state.case_logs[res_name] = resident_df
                        total_imported += 1
        return total_imported

    # File upload section
    with st.expander("📤 Upload Case Logs", expanded=True):
        st.markdown("""
**How to export from ACGME:**
1. Log into [ACGME ADS](https://apps.acgme.org/connect/login)
2. Go to **Case Logs** → **Reports** → **Resident Minimum Defined Categories**
3. Select resident(s) and export to **PDF** or Excel
4. Upload below

Supports PDF and Excel formats.
        """)

        uploaded_file = st.file_uploader(
            "Upload ACGME Case Log Export (PDF/Excel/CSV)",
            type=["pdf", "xlsx", "xls", "csv"],
            key="case_log_upload"
        )

        # Show OCR availability
        if HAS_OCR:
            st.caption("OCR enabled for image-based PDFs")
        else:
            st.caption("OCR not available - text-based PDFs only")

        # Debug mode checkbox
        show_pdf_debug = st.checkbox("Show PDF diagnostic info", value=False, key="pdf_debug_mode")

        if uploaded_file:
            try:
                df = None
                if uploaded_file.name.lower().endswith('.pdf'):
                    if HAS_PDFPLUMBER:
                        file_buffer = io.BytesIO(uploaded_file.read())

                        # Show diagnostic info if requested
                        if show_pdf_debug:
                            file_buffer.seek(0)
                            st.markdown("### PDF Diagnostic Info")
                            st.write(f"**OCR Available:** {HAS_OCR}")
                            with pdfplumber.open(file_buffer) as pdf:
                                st.write(f"**Pages:** {len(pdf.pages)}")
                                has_any_text = False
                                for i, page in enumerate(pdf.pages):
                                    st.markdown(f"#### Page {i+1}")

                                    # Show extracted text
                                    text = page.extract_text()
                                    if text:
                                        has_any_text = True
                                        st.markdown("**Extracted Text (first 2000 chars):**")
                                        st.code(text[:2000])
                                    else:
                                        st.warning("No text extracted from this page")

                                    # Show extracted tables
                                    tables = page.extract_tables()
                                    if tables:
                                        st.markdown(f"**Tables found:** {len(tables)}")
                                        for j, table in enumerate(tables):
                                            st.markdown(f"Table {j+1} ({len(table)} rows):")
                                            if table:
                                                # Show first 10 rows
                                                preview = table[:10]
                                                st.dataframe(pd.DataFrame(preview), use_container_width=True)
                                    else:
                                        st.warning("No tables extracted from this page")

                                if not has_any_text:
                                    st.markdown("---")
                                    st.info("This appears to be an **image-based PDF**. OCR will be used to extract text.")

                            st.markdown("---")
                            file_buffer.seek(0)

                        # Parse the PDF (will auto-fallback to OCR if needed)
                        with st.spinner("Processing PDF (may use OCR for image-based PDFs)..."):
                            df = parse_acgme_pdf(file_buffer, debug=show_pdf_debug)

                        if df is None or df.empty:
                            if not show_pdf_debug:
                                st.error("Could not extract data from PDF.")
                                if HAS_OCR:
                                    st.info("OCR was attempted but could not parse the content. Try Excel format.")
                                else:
                                    st.info("**Enable 'Show PDF diagnostic info' above** to see details.")
                    else:
                        st.error("PDF support not available. Please upload Excel format.")
                elif uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)

                if df is not None and not df.empty:
                    st.success(f"✅ Loaded file with {len(df)} rows")

                    # Check if it's summary format
                    if is_summary_format(df):
                        st.info("📊 Detected ACGME summary format (category totals)")

                        # Show preview
                        st.markdown("**Preview:**")
                        st.dataframe(df.head(10), use_container_width=True)

                        if st.button("Import Summary Data"):
                            imported = import_summary_file(df, uploaded_file.name)
                            if imported:
                                save_case_logs_to_cache(st.session_state.case_logs)
                                st.success(f"✅ Imported data for: {', '.join(imported)}")
                                st.rerun()
                            else:
                                st.error("Could not parse resident data from file. Check the format.")
                    else:
                        # Individual case format - need column mapping
                        st.info("📋 Detected individual case format")
                        df.columns = df.columns.str.strip()
                        possible_resident_cols = [c for c in df.columns if any(x in c.lower() for x in ['resident', 'name', 'trainee'])]
                        possible_cpt_cols = [c for c in df.columns if any(x in c.lower() for x in ['cpt', 'code', 'procedure'])]

                        resident_col = st.selectbox("Select Resident column:", options=df.columns.tolist(),
                                                   index=df.columns.tolist().index(possible_resident_cols[0]) if possible_resident_cols else 0)
                        cpt_col = st.selectbox("Select CPT Code column:", options=df.columns.tolist(),
                                               index=df.columns.tolist().index(possible_cpt_cols[0]) if possible_cpt_cols else 0)
                        role_col = st.selectbox("Select Role column (optional):", options=["(None)"] + df.columns.tolist())

                        if st.button("Import Cases"):
                            for resident in df[resident_col].unique():
                                if pd.isna(resident) or str(resident).strip() == "":
                                    continue
                                resident_df = df[df[resident_col] == resident].copy()
                                if resident not in st.session_state.case_logs:
                                    st.session_state.case_logs[resident] = resident_df
                                else:
                                    if isinstance(st.session_state.case_logs[resident], pd.DataFrame):
                                        st.session_state.case_logs[resident] = pd.concat(
                                            [st.session_state.case_logs[resident], resident_df]
                                        ).drop_duplicates()
                            save_case_logs_to_cache(st.session_state.case_logs)
                            st.success(f"✅ Imported cases for {len(df[resident_col].unique())} residents")
                            st.rerun()

            except Exception as e:
                st.error(f"Error reading file: {e}")

    # Dashboard display
    if st.session_state.case_logs:
        st.markdown("---")
        st.markdown("### 📈 Case Log Progress by Resident")

        # Resident selector
        all_residents = sorted(st.session_state.case_logs.keys())
        selected_resident = st.selectbox("Select Resident:", options=["(All Residents)"] + all_residents,
                                         key="case_log_resident_view")

        def render_progress_bar(current, minimum, label):
            """Render a progress bar with color coding."""
            if minimum == 0:
                pct = 100 if current > 0 else 0
            else:
                pct = min(100, (current / minimum) * 100)

            color = "#28a745" if pct >= 100 else "#ffc107" if pct >= 50 else "#dc3545"

            st.markdown(f"""
            <div style="margin-bottom: 8px;">
                <div style="display: flex; justify-content: space-between; margin-bottom: 2px;">
                    <span>{label}</span>
                    <span><b>{current}</b> / {minimum}</span>
                </div>
                <div style="background-color: #e0e0e0; border-radius: 4px; height: 20px;">
                    <div style="background-color: {color}; width: {pct}%; height: 100%; border-radius: 4px;"></div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        def get_resident_progress(res_data):
            """Get progress data from either summary dict or DataFrame."""
            if isinstance(res_data, dict):
                # Summary format: {category: {"total": N, "minimum": M, "subcategories": {...}}}
                return res_data
            elif isinstance(res_data, pd.DataFrame):
                # Individual cases format - count by CPT codes
                counts = {cat: {"total": 0, "minimum": ACGME_CATEGORIES[cat]["minimum"], "subcategories": {}} for cat in ACGME_CATEGORIES}
                cpt_col = None
                for c in res_data.columns:
                    if 'cpt' in c.lower() or 'code' in c.lower():
                        cpt_col = c
                        break
                if cpt_col:
                    for _, row in res_data.iterrows():
                        cpt = str(row.get(cpt_col, "")).strip()
                        if cpt in CPT_TO_CATEGORY:
                            cat, subcat = CPT_TO_CATEGORY[cpt]
                            if cat in counts:
                                counts[cat]["total"] += 1
                return counts
            return {}

        if selected_resident == "(All Residents)":
            # Summary view for all residents
            summary_data = []
            for res in all_residents:
                res_data = st.session_state.case_logs[res]
                progress = get_resident_progress(res_data)
                total_counted = sum(cat_data.get("total", 0) for cat_data in progress.values())

                # Get PGY level from roster
                roster_match = st.session_state.roster_table[
                    st.session_state.roster_table["Resident"].str.strip() == res.strip()
                ]
                pgy = roster_match["PGY"].iloc[0] if not roster_match.empty else "?"

                # Calculate completion percentage
                pct = (total_counted / ACGME_TOTAL_MINIMUM * 100) if ACGME_TOTAL_MINIMUM > 0 else 0

                summary_data.append({
                    "Resident": res,
                    "PGY": pgy,
                    "Total Cases": total_counted,
                    "Progress": f"{total_counted}/{ACGME_TOTAL_MINIMUM} ({pct:.0f}%)"
                })

            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

            # Bar chart of total cases by resident
            if summary_data:
                chart_df = pd.DataFrame(summary_data)
                chart = alt.Chart(chart_df).mark_bar().encode(
                    x=alt.X("Resident:N", sort="-y"),
                    y=alt.Y("Total Cases:Q"),
                    color=alt.Color("PGY:N")
                ).properties(height=300, title="Total Cases by Resident")
                st.altair_chart(chart, use_container_width=True)

        else:
            # Individual resident view
            res_data = st.session_state.case_logs[selected_resident]
            progress = get_resident_progress(res_data)

            st.markdown(f"### {selected_resident}")

            # Total progress
            total_counted = sum(cat_data.get("total", 0) for cat_data in progress.values())
            st.markdown("#### Overall Progress")
            render_progress_bar(total_counted, ACGME_TOTAL_MINIMUM, "Total Major Cases")

            # Category breakdown
            st.markdown("#### Progress by Category")

            for cat, cat_info in ACGME_CATEGORIES.items():
                minimum = cat_info["minimum"]
                # Get current from progress data (either from summary or calculated)
                cat_progress = progress.get(cat, {})
                current = cat_progress.get("total", 0)
                # Use minimum from import if available, otherwise from ACGME_CATEGORIES
                minimum_display = cat_progress.get("minimum", minimum)

                with st.expander(f"{cat} ({current}/{minimum_display})", expanded=(current < minimum_display)):
                    render_progress_bar(current, minimum_display, cat)

                    # Subcategories from either source
                    subcats = cat_progress.get("subcategories", {})
                    expected_subcats = cat_info.get("subcategories", {})

                    # Show subcategories from progress data
                    for subcat, sub_data in subcats.items():
                        if isinstance(sub_data, dict):
                            sub_cur = sub_data.get("count", 0)
                            sub_min = sub_data.get("minimum", expected_subcats.get(subcat, {}).get("minimum", 0))
                        else:
                            sub_cur = sub_data
                            sub_min = expected_subcats.get(subcat, {}).get("minimum", 0)
                        if sub_min > 0 or sub_cur > 0:
                            render_progress_bar(sub_cur, sub_min, f"  └─ {subcat}")

            # Clear data button
            if st.button(f"🗑️ Clear data for {selected_resident}", key="clear_resident_data"):
                del st.session_state.case_logs[selected_resident]
                save_case_logs_to_cache(st.session_state.case_logs)
                st.success(f"Cleared data for {selected_resident}")
                st.rerun()

    else:
        st.info("No case log data loaded. Drop an ACGME summary export (ResMinimumDefCat) into the watch folder.")

# Weekly Schedule tab
with tabs[5]:
    st.markdown("## 🗓️ Weekly OR & Clinic Schedule")
    st.caption("Manage the weekly surgery schedule and auto-assign residents to cases.")

    # Initialize weekly schedule storage
    if "weekly_schedule" not in st.session_state:
        st.session_state.weekly_schedule = []  # List of case dicts

    # Date selector
    col_date1, col_date2 = st.columns(2)
    with col_date1:
        schedule_week_start = st.date_input("Week starting:", value=date.today() - timedelta(days=date.today().weekday()),
                                            key="schedule_week_start")

    st.markdown("---")

    # Add case section
    with st.expander("➕ Add Surgery / Clinic", expanded=True):
        add_cols = st.columns([1, 1, 1, 2, 1])

        with add_cols[0]:
            case_date = st.date_input("Date:", value=schedule_week_start, key="new_case_date")
        with add_cols[1]:
            case_time = st.time_input("Time:", value=None, key="new_case_time")
        with add_cols[2]:
            case_type = st.selectbox("Type:", ["OR Case", "Clinic"], key="new_case_type")

        # Attending selection grouped by team
        all_attending_names = list(ALL_ATTENDINGS.keys())
        with add_cols[3]:
            case_attending = st.selectbox("Attending:", options=all_attending_names, key="new_case_attending")

        with add_cols[4]:
            attending_team = ALL_ATTENDINGS.get(case_attending, "Other")
            st.text_input("Team:", value=attending_team, disabled=True, key="new_case_team_display")

        # Case details
        detail_cols = st.columns([2, 1, 2])
        with detail_cols[0]:
            case_procedure = st.text_input("Procedure/Description:", key="new_case_procedure")
        with detail_cols[1]:
            case_cpt = st.text_input("CPT Code (optional):", key="new_case_cpt")
        with detail_cols[2]:
            case_location = st.text_input("Location/Room:", key="new_case_location")

        # Resident assignment
        st.markdown("**Resident Assignment:**")
        assign_cols = st.columns([1, 2, 2])

        # Get residents on the relevant team this block
        def get_residents_on_team(team_name):
            """Get residents currently assigned to a team based on yearly schedule."""
            residents = []
            schedule_df = st.session_state.get("schedule_df", None)
            if schedule_df is None:
                return list(st.session_state.roster_table["Resident"].dropna().unique())

            # Find current block
            blocks = build_blocks(ay_start, 13)
            current_block_idx = None
            for i, (s, e) in enumerate(blocks):
                if s <= case_date <= e:
                    current_block_idx = i
                    break

            if current_block_idx is not None:
                hdr = hdr_for_block(blocks[current_block_idx][0], blocks[current_block_idx][1])
                if hdr in schedule_df.columns:
                    team_norm = norm_label(team_name)
                    for res in schedule_df.index:
                        if norm_label(schedule_df.loc[res, hdr]) == team_norm:
                            residents.append(res)

            # If no residents found on team, return all
            if not residents:
                return list(st.session_state.roster_table["Resident"].dropna().unique())

            return residents

        team_residents = get_residents_on_team(attending_team)
        all_residents_list = list(st.session_state.roster_table["Resident"].dropna().unique())

        with assign_cols[0]:
            auto_assign = st.checkbox("Auto-assign", value=True, key="new_case_auto_assign",
                                     help="Automatically assign based on PGY level and case needs")

        with assign_cols[1]:
            primary_resident = st.selectbox("Primary Resident:",
                                           options=["(Auto)"] + team_residents if auto_assign else all_residents_list,
                                           key="new_case_primary_resident")

        with assign_cols[2]:
            secondary_resident = st.selectbox("Secondary Resident (optional):",
                                             options=["(None)"] + all_residents_list,
                                             key="new_case_secondary_resident")

        if st.button("Add to Schedule", type="primary"):
            new_case = {
                "date": case_date.isoformat(),
                "time": case_time.isoformat() if case_time else "",
                "type": case_type,
                "attending": case_attending,
                "team": attending_team,
                "procedure": case_procedure,
                "cpt": case_cpt,
                "location": case_location,
                "primary_resident": primary_resident if primary_resident != "(Auto)" else "",
                "secondary_resident": secondary_resident if secondary_resident != "(None)" else "",
                "auto_assign": auto_assign
            }
            st.session_state.weekly_schedule.append(new_case)
            st.success(f"Added: {case_procedure} with {case_attending}")
            st.rerun()

    # Upload Epic export option
    with st.expander("📤 Upload Schedule (Epic Export)"):
        st.markdown("""
**How to export from Epic Reporting Workbench:**
1. Run your OR Schedule report for the week
2. Export as Excel
3. Upload below
        """)

        epic_upload = st.file_uploader("Upload Epic Schedule Export:", type=["xlsx", "xls", "csv"],
                                       key="epic_schedule_upload")

        if epic_upload:
            try:
                if epic_upload.name.endswith('.csv'):
                    epic_df = pd.read_csv(epic_upload)
                else:
                    epic_df = pd.read_excel(epic_upload)

                st.success(f"✅ Loaded {len(epic_df)} rows")
                st.dataframe(epic_df.head(), use_container_width=True)

                # Column mapping
                st.markdown("**Map columns:**")
                map_cols = st.columns(4)
                with map_cols[0]:
                    epic_date_col = st.selectbox("Date column:", epic_df.columns.tolist(), key="epic_date_col")
                with map_cols[1]:
                    epic_proc_col = st.selectbox("Procedure column:", epic_df.columns.tolist(), key="epic_proc_col")
                with map_cols[2]:
                    epic_attending_col = st.selectbox("Attending column:", epic_df.columns.tolist(), key="epic_attending_col")
                with map_cols[3]:
                    epic_loc_col = st.selectbox("Location column:", ["(None)"] + epic_df.columns.tolist(), key="epic_loc_col")

                if st.button("Import Schedule"):
                    imported = 0
                    for _, row in epic_df.iterrows():
                        try:
                            case_date_val = pd.to_datetime(row[epic_date_col]).date()
                            attending_name = str(row[epic_attending_col]).strip()

                            # Try to match attending
                            matched_attending = None
                            for att in ALL_ATTENDINGS:
                                if att.lower() in attending_name.lower() or attending_name.lower() in att.lower():
                                    matched_attending = att
                                    break

                            new_case = {
                                "date": case_date_val.isoformat(),
                                "time": "",
                                "type": "OR Case",
                                "attending": matched_attending or attending_name,
                                "team": ALL_ATTENDINGS.get(matched_attending, "Other") if matched_attending else "Other",
                                "procedure": str(row[epic_proc_col]),
                                "cpt": "",
                                "location": str(row[epic_loc_col]) if epic_loc_col != "(None)" else "",
                                "primary_resident": "",
                                "secondary_resident": "",
                                "auto_assign": True
                            }
                            st.session_state.weekly_schedule.append(new_case)
                            imported += 1
                        except Exception:
                            continue

                    st.success(f"Imported {imported} cases")
                    st.rerun()

            except Exception as e:
                st.error(f"Error reading file: {e}")

    # Display current schedule
    st.markdown("---")
    st.markdown("### 📋 This Week's Schedule")

    # Filter to current week
    week_end = schedule_week_start + timedelta(days=6)
    week_cases = [c for c in st.session_state.weekly_schedule
                  if schedule_week_start <= date.fromisoformat(c["date"]) <= week_end]

    if week_cases:
        # Auto-assign residents to cases that need it
        def auto_assign_resident_to_case(case_dict):
            """Auto-assign resident based on PGY level and case needs."""
            if case_dict.get("primary_resident") and case_dict["primary_resident"] != "(Auto)":
                return case_dict["primary_resident"], case_dict.get("secondary_resident", "")

            # Get residents on the team
            team_residents = get_residents_on_team(case_dict["team"])
            if not team_residents:
                return "", ""

            # Get case logs to check who needs what
            cpt = case_dict.get("cpt", "")
            case_category = None
            if cpt and cpt in CPT_TO_CATEGORY:
                case_category = CPT_TO_CATEGORY[cpt][0]

            # Score residents: prioritize higher PGY who still need the case type
            def score_resident(res):
                # Get PGY level
                roster_match = st.session_state.roster_table[
                    st.session_state.roster_table["Resident"].str.strip() == res.strip()
                ]
                if roster_match.empty:
                    return (0, 0)

                pgy_str = roster_match["PGY"].iloc[0]
                pgy_num = int(pgy_str.replace("PGY-", "")) if "PGY-" in str(pgy_str) else 0

                # Check if they need this case type
                need_score = 0
                if case_category and res in st.session_state.case_logs:
                    counts = count_cases_by_category(st.session_state.case_logs[res])
                    if case_category in counts:
                        minimum = ACGME_CATEGORIES.get(case_category, {}).get("minimum", 0)
                        current = counts[case_category]["total"]
                        if current < minimum:
                            need_score = minimum - current  # Higher need = higher score

                return (pgy_num, need_score)

            # Sort by PGY (descending), then by need (descending)
            scored = [(res, score_resident(res)) for res in team_residents]
            scored.sort(key=lambda x: (x[1][0], x[1][1]), reverse=True)

            primary = scored[0][0] if scored else ""

            # For Teaching Assist (PGY-4+), can assign a junior too
            secondary = ""
            if len(scored) > 1 and scored[0][1][0] >= 4:  # PGY-4 or higher
                # Find a junior who needs the case
                juniors = [s for s in scored[1:] if s[1][0] < scored[0][1][0]]
                if juniors:
                    secondary = juniors[0][0]

            return primary, secondary

        # Group by date
        from collections import defaultdict
        by_date = defaultdict(list)
        for case in week_cases:
            by_date[case["date"]].append(case)

        for case_date_str in sorted(by_date.keys()):
            case_date_obj = date.fromisoformat(case_date_str)
            day_name = case_date_obj.strftime("%A, %B %d")
            st.markdown(f"#### {day_name}")

            day_cases = by_date[case_date_str]
            for i, case in enumerate(day_cases):
                # Auto-assign if needed
                if case.get("auto_assign") and not case.get("primary_resident"):
                    primary, secondary = auto_assign_resident_to_case(case)
                    case["primary_resident"] = primary
                    case["secondary_resident"] = secondary

                # Display case
                time_str = case.get("time", "")
                if time_str:
                    time_display = f"**{time_str[:5]}**"
                else:
                    time_display = ""

                col1, col2, col3, col4 = st.columns([1, 2, 2, 1])

                with col1:
                    st.write(time_display)
                    st.caption(case.get("type", "OR"))

                with col2:
                    st.write(f"**{case.get('procedure', 'TBD')}**")
                    st.caption(f"{case.get('attending', '')} ({case.get('team', '')})")

                with col3:
                    primary = case.get("primary_resident", "")
                    secondary = case.get("secondary_resident", "")
                    if primary:
                        st.write(f"👨‍⚕️ **{primary}**")
                        if secondary:
                            st.caption(f"+ {secondary} (assist)")
                    else:
                        st.write("⚠️ *Unassigned*")

                with col4:
                    if st.button("❌", key=f"del_case_{case_date_str}_{i}"):
                        st.session_state.weekly_schedule.remove(case)
                        st.rerun()

                st.markdown("---")

        # Summary stats
        st.markdown("### 📊 Week Summary")
        total_or = len([c for c in week_cases if c["type"] == "OR Case"])
        total_clinic = len([c for c in week_cases if c["type"] == "Clinic"])

        sum_cols = st.columns(3)
        with sum_cols[0]:
            st.metric("OR Cases", total_or)
        with sum_cols[1]:
            st.metric("Clinic Sessions", total_clinic)
        with sum_cols[2]:
            assigned = len([c for c in week_cases if c.get("primary_resident")])
            st.metric("Assigned", f"{assigned}/{len(week_cases)}")

    else:
        st.info("No cases scheduled for this week. Add cases above or upload an Epic export.")
